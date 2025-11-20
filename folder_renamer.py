import os
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import datetime
import json
import tkinter.simpledialog
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ModernFolderRenamer:
    def __init__(self, root):
        self.root = root
        self.root.title("Folder Manager - Kozen v2.9.2")
        self.root.geometry("1200x750")
        self.root.configure(bg='#f8f9fa')
        self.root.minsize(1000, 600)
        
        # Центрирование окна
        self.center_window()
        
        # Загрузка конфигурации атак
        self.config_file = "attack_config.json"
        self.load_attack_config()
        
        # Данные для отчёта
        self.shooting_report_data = []
        
        # Стили
        self.setup_styles()
        self.setup_ui()
        
    def center_window(self):
        """Центрирование окна на экране"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Современная цветовая схема
        self.colors = {
            'primary': '#4f46e5',
            'primary_light': '#6366f1',
            'secondary': '#64748b',
            'success': '#10b981',
            'warning': '#f59e0b',
            'error': '#ef4444',
            'background': '#f8f9fa',
            'surface': '#ffffff',
            'text_primary': '#1e293b',
            'text_secondary': '#64748b',
            'border': '#e2e8f0'
        }
        
        # Конфигурация стилей
        self.style.configure('TFrame', background=self.colors['background'])
        self.style.configure('TLabel', background=self.colors['background'], foreground=self.colors['text_primary'])
        self.style.configure('TButton', font=('Segoe UI', 9), borderwidth=0, focuscolor='none')
        self.style.configure('Rounded.TButton', 
                           background=self.colors['primary'],
                           foreground='white',
                           borderwidth=0,
                           focuscolor='none',
                           relief='flat',
                           padding=(15, 8))
        self.style.map('Rounded.TButton',
                      background=[('active', self.colors['primary_light']),
                                ('pressed', self.colors['primary'])])
        
        self.style.configure('Secondary.TButton',
                           background=self.colors['surface'],
                           foreground=self.colors['text_primary'],
                           borderwidth=1,
                           relief='solid')
        self.style.map('Secondary.TButton',
                      background=[('active', self.colors['border'])])

        self.style.configure('Success.TButton',
                           background=self.colors['success'],
                           foreground='white')
        self.style.map('Success.TButton',
                      background=[('active', '#34d399')])

        self.style.configure('Warning.TButton',
                           background=self.colors['warning'],
                           foreground='white')
        self.style.map('Warning.TButton',
                      background=[('active', '#fbbf24')])

        self.style.configure('Rounded.TFrame', 
                           background=self.colors['surface'],
                           relief='solid',
                           borderwidth=1)
        
    def load_attack_config(self):
        """Загрузка конфигурации атак из JSON файла"""
        default_config = {
            "02 2D Mask": {"kozen 10": (91, 170), "kozen 12": (171, 250)},
            "03 2D Mask": {"kozen 10": (251, 346), "kozen 12": (347, 442)},
            "04 2D Mask": {"kozen 10": (443, 458), "kozen 12": (459, 474)},
            "05 2D Mask": {"kozen 10": (475, 514), "kozen 12": (515, 554)},
            "06 2D Mask dev 1": {"kozen 10": (555, 594)},
            "07 2D Mask dev 2": {"kozen 12": (595, 634)},
            "08 3D Mask dev 1": {"kozen 10": (635, 733)},
            "09 3D Mask dev 2": {"kozen 12": (734, 832)},
            "10 Indoors": {"kozen 10": (833, 848), "kozen 12": (975, 990)},
            "11 Indoors. With attributes": {"kozen 10": (849, 876), "kozen 12": (991, 1018)},
            "12 Indoors. Backlight": {"kozen 10": (877, 890), "kozen 12": (1019, 1032)},
            "13 Indoors. Insufficient lighting": {"kozen 10": (891, 918), "kozen 12": (1033, 1060)},
            "14 Indoors. Behind transparent glass": {"kozen 10": (919, 932), "kozen 12": (1061, 1074)},
            "15 Outside": {"kozen 10": (933, 974), "kozen 12": (1075, 1116)}
        }
        
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    loaded_config = json.load(f)
                    # Конвертируем списки обратно в кортежи и исправляем опечатки в названиях устройств
                    self.attack_ranges = {}
                    for attack, devices in loaded_config.items():
                        self.attack_ranges[attack] = {}
                        for device, range_tuple in devices.items():
                            # Исправляем опечатки в названиях устройств
                            normalized_device = device.replace("kozen 101", "kozen 10").replace("kozen 121", "kozen 12")
                            if isinstance(range_tuple, list):
                                self.attack_ranges[attack][normalized_device] = tuple(range_tuple)
                            else:
                                self.attack_ranges[attack][normalized_device] = range_tuple
            else:
                self.attack_ranges = default_config
                self.save_attack_config()
        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")
            self.attack_ranges = default_config
            self.save_attack_config()
    
    def save_attack_config(self):
        """Сохранение конфигурации атак в JSON файл"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.attack_ranges, f, indent=4, ensure_ascii=False)
        except Exception as e:
            self.log(f"Ошибка сохранения конфигурации: {str(e)}", "ERROR")
    
    def create_rounded_frame(self, parent, **kwargs):
        """Создание фрейма с скруглёнными краями"""
        frame = tk.Frame(parent, 
                        bg=self.colors['surface'],
                        relief='solid',
                        bd=1,
                        **kwargs)
        return frame
    
    def setup_ui(self):
        # Заголовок
        header_frame = self.create_rounded_frame(self.root)
        header_frame.pack(fill="x", padx=15, pady=10)
        
        title_label = tk.Label(header_frame, 
                              text="📁 Folder Manager - Kozen", 
                              font=("Segoe UI", 18, "bold"),
                              bg=self.colors['surface'], 
                              fg=self.colors['text_primary'],
                              pady=12)
        title_label.pack()
        
        # Основной контейнер с вкладками
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=15, pady=8)
        
        # Вкладка основного функционала
        main_tab = self.create_rounded_frame(notebook)
        notebook.add(main_tab, text="🔄 Основные функции")
        
        # Вкладка проверки
        check_tab = self.create_rounded_frame(notebook)
        notebook.add(check_tab, text="🔍 Проверка")
        
        # Вкладка настроек атак
        settings_tab = self.create_rounded_frame(notebook)
        notebook.add(settings_tab, text="⚙️ Настройки атак")
        
        self.setup_main_tab(main_tab)
        self.setup_check_tab(check_tab)
        self.setup_settings_tab(settings_tab)
    
    def setup_main_tab(self, parent):
        # Создаем разделяемый фрейм для левой (настройки) и правой (логи) части
        main_paned = ttk.PanedWindow(parent, orient=tk.HORIZONTAL)
        main_paned.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Левая часть - настройки
        left_frame = self.create_rounded_frame(main_paned)
        main_paned.add(left_frame, weight=1)
        
        # Правая часть - логи
        right_frame = self.create_rounded_frame(main_paned)
        main_paned.add(right_frame, weight=1)
        
        # Настройка левой части - элементы управления
        # Фрейм для выбора папок
        folder_frame = self.create_rounded_frame(left_frame)
        folder_frame.pack(fill="x", padx=10, pady=8)
        
        # Исходная папка
        tk.Label(folder_frame, text="📂 Исходная папка:", 
                font=("Segoe UI", 9, "bold"),
                bg=self.colors['surface']).grid(row=0, column=0, sticky="w", pady=(12, 4), padx=12)
        
        input_frame1 = tk.Frame(folder_frame, bg=self.colors['surface'])
        input_frame1.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))
        input_frame1.columnconfigure(0, weight=1)
        
        self.source_entry = ttk.Entry(input_frame1, font=("Segoe UI", 9))
        self.source_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        
        ttk.Button(input_frame1, text="Обзор", 
                  command=self.browse_source, style="Secondary.TButton").grid(row=0, column=1)
        
        # Информация о количестве папок в исходной папке
        self.source_info_label = tk.Label(folder_frame, text="", font=("Segoe UI", 8),
                                         bg=self.colors['surface'], fg=self.colors['text_secondary'])
        self.source_info_label.grid(row=2, column=0, sticky="w", pady=(0, 8), padx=12)
        
        # Папка назначения
        tk.Label(folder_frame, text="📁 Папка назначения:", 
                font=("Segoe UI", 9, "bold"),
                bg=self.colors['surface']).grid(row=3, column=0, sticky="w", pady=(8, 4), padx=12)
        
        input_frame2 = tk.Frame(folder_frame, bg=self.colors['surface'])
        input_frame2.grid(row=4, column=0, sticky="ew", padx=12, pady=(0, 12))
        input_frame2.columnconfigure(0, weight=1)
        
        self.dest_entry = ttk.Entry(input_frame2, font=("Segoe UI", 9))
        self.dest_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        
        ttk.Button(input_frame2, text="Обзор", 
                  command=self.browse_dest, style="Secondary.TButton").grid(row=0, column=1)
        
        folder_frame.columnconfigure(0, weight=1)
        
        # Обновляем информацию при изменении пути
        self.source_entry.bind('<KeyRelease>', self.update_source_info)
        self.source_entry.bind('<FocusOut>', self.update_source_info)
        
        # Фрейм для настроек
        settings_frame = self.create_rounded_frame(left_frame)
        settings_frame.pack(fill="x", padx=10, pady=8)
        
        # Устройство
        tk.Label(settings_frame, text="📱 Устройство:", 
                font=("Segoe UI", 9, "bold"),
                bg=self.colors['surface']).grid(row=0, column=0, sticky="w", pady=(12, 8), padx=12)
        
        self.device_var = tk.StringVar(value="все")
        device_frame = tk.Frame(settings_frame, bg=self.colors['surface'])
        device_frame.grid(row=0, column=1, sticky="w", pady=(12, 8), padx=12)
        
        ttk.Radiobutton(device_frame, text="Все", variable=self.device_var, 
                       value="все", command=self.update_range_info).pack(side="left", padx=(0, 15))
        ttk.Radiobutton(device_frame, text="Kozen 10", variable=self.device_var, 
                       value="kozen 10", command=self.update_range_info).pack(side="left", padx=(0, 15))
        ttk.Radiobutton(device_frame, text="Kozen 12", variable=self.device_var, 
                       value="kozen 12", command=self.update_range_info).pack(side="left")
        
        # Атака
        tk.Label(settings_frame, text="🎯 Тип атаки:", 
                font=("Segoe UI", 9, "bold"),
                bg=self.colors['surface']).grid(row=1, column=0, sticky="w", pady=8, padx=12)
        
        self.attack_var = tk.StringVar(value="02 2D Mask")
        self.attack_combo = ttk.Combobox(settings_frame, textvariable=self.attack_var, 
                                       values=list(self.attack_ranges.keys()), 
                                       state="readonly", font=("Segoe UI", 9))
        self.attack_combo.grid(row=1, column=1, sticky="w", pady=8, padx=12)
        self.attack_combo.bind("<<ComboboxSelected>>", self.update_range_info)
        
        # Чекбокс проверки содержимого
        self.check_content_var = tk.BooleanVar(value=False)
        check_frame = tk.Frame(settings_frame, bg=self.colors['surface'])
        check_frame.grid(row=2, column=0, columnspan=2, sticky="w", pady=8, padx=12)
        
        ttk.Checkbutton(check_frame, text="🔍 Проверять содержимое папок (3 папки + BestShot файл)", 
                       variable=self.check_content_var, style="TCheckbutton").pack(side="left")
        
        # Информация о диапазоне
        self.range_info = tk.Label(settings_frame, text="", font=("Segoe UI", 9),
                                  bg=self.colors['surface'], fg=self.colors['primary'],
                                  pady=8)
        self.range_info.grid(row=3, column=0, columnspan=2, sticky="w", padx=12)
        
        settings_frame.columnconfigure(1, weight=1)
        
        # Фрейм для замены папок
        replace_frame = self.create_rounded_frame(left_frame)
        replace_frame.pack(fill="x", padx=10, pady=8)
        
        tk.Label(replace_frame, text="🔧 Замена отдельных папок", 
                font=("Segoe UI", 9, "bold"),
                bg=self.colors['surface']).pack(anchor="w", pady=(12, 8), padx=12)
        
        tk.Label(replace_frame, text="Номера папок для замены:", 
                font=("Segoe UI", 9),
                bg=self.colors['surface']).pack(anchor="w", padx=12)
        
        input_frame = tk.Frame(replace_frame, bg=self.colors['surface'])
        input_frame.pack(fill="x", padx=12, pady=8)
        input_frame.columnconfigure(0, weight=1)
        
        self.replace_entry = ttk.Entry(input_frame, font=("Segoe UI", 9))
        self.replace_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        
        tk.Label(replace_frame, text="Пример: 522, 530-532,528", 
                font=("Segoe UI", 8),
                bg=self.colors['surface'],
                fg=self.colors['text_secondary']).pack(anchor="w", padx=12, pady=(0, 12))
        
        # Кнопки выполнения - ВЕРТИКАЛЬНО для маленьких экранов
        button_frame = tk.Frame(left_frame, bg=self.colors['background'])
        button_frame.pack(fill="x", padx=10, pady=10)
        
        # Контейнер для кнопок - вертикальное расположение
        btn_container = tk.Frame(button_frame, bg=self.colors['background'])
        btn_container.pack(fill="x")
        
        # Кнопки одинакового размера, расположенные вертикально
        self.rename_btn = ttk.Button(btn_container, text="🚀 Выполнить переименование", 
                                   command=self.execute_renaming, 
                                   style="Rounded.TButton")
        self.rename_btn.pack(fill="x", pady=2)
        
        self.replace_btn = ttk.Button(btn_container, text="🔄 Выполнить замену", 
                                    command=self.execute_replacement, 
                                    style="Warning.TButton")
        self.replace_btn.pack(fill="x", pady=2)
        
        self.update_range_info()
        
        # Настройка правой части - логи
        log_header_frame = tk.Frame(right_frame, bg=self.colors['surface'])
        log_header_frame.pack(fill="x", padx=12, pady=(12, 8))
        
        tk.Label(log_header_frame, text="📋 Основные логи выполнения", 
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['surface']).pack(side="left")
        
        # Кнопка очистки логов в заголовке
        ttk.Button(log_header_frame, text="🧹 Очистить логи", 
                  command=self.clear_logs, style="Secondary.TButton").pack(side="right")
        
        # Фрейм для логов с прокруткой
        log_container = tk.Frame(right_frame, bg=self.colors['surface'])
        log_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        
        self.log_text = scrolledtext.ScrolledText(log_container, height=20, font=("Consolas", 8),
                                                 bg='#1e293b', fg='#e2e8f0', 
                                                 insertbackground='white',
                                                 relief='flat',
                                                 padx=8, pady=8)
        self.log_text.pack(fill="both", expand=True)
        
        # Настройка тегов для цветного текста
        self.log_text.tag_config("SUCCESS", foreground="#10b981")
        self.log_text.tag_config("WARNING", foreground="#f59e0b")
        self.log_text.tag_config("ERROR", foreground="#ef4444")
        self.log_text.tag_config("INFO", foreground="#e2e8f0")
        self.log_text.tag_config("CRITICAL", foreground="#ff0000", background="#330000")
        self.log_text.tag_config("HEADER", foreground="#93c5fd", font=("Consolas", 8, "bold"))
        self.log_text.tag_config("DETAIL", foreground="#94a3b8")
    
    def setup_check_tab(self, parent):
        paned_window = ttk.PanedWindow(parent, orient=tk.HORIZONTAL)
        paned_window.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Левая часть - элементы управления
        left_frame = self.create_rounded_frame(paned_window)
        paned_window.add(left_frame, weight=1)
        
        # Правая часть - логи
        right_frame = self.create_rounded_frame(paned_window)
        paned_window.add(right_frame, weight=2)
        
        # Настройка левой части
        attack_check_frame = self.create_rounded_frame(left_frame)
        attack_check_frame.pack(fill="x", padx=10, pady=8)
        
        tk.Label(attack_check_frame, text="🎯 Проверка атаки", 
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['surface']).pack(anchor="w", pady=(12, 8), padx=12)
        
        input_frame1 = tk.Frame(attack_check_frame, bg=self.colors['surface'])
        input_frame1.pack(fill="x", padx=12, pady=8)
        
        tk.Label(input_frame1, text="Папка атаки:", 
                font=("Segoe UI", 9),
                bg=self.colors['surface']).grid(row=0, column=0, sticky="w")
        
        self.attack_check_entry = ttk.Entry(input_frame1, font=("Segoe UI", 9))
        self.attack_check_entry.grid(row=0, column=1, sticky="ew", padx=8)
        
        ttk.Button(input_frame1, text="Обзор", 
                  command=lambda: self.browse_folder(self.attack_check_entry),
                  style="Secondary.TButton").grid(row=0, column=2, padx=(5, 0))
        
        input_frame1.columnconfigure(1, weight=1)
        
        button_frame1 = tk.Frame(attack_check_frame, bg=self.colors['surface'])
        button_frame1.pack(fill="x", padx=12, pady=8)
        
        ttk.Button(button_frame1, text="🔍 Проверить атаку", 
                  command=self.check_attack, 
                  style="Rounded.TButton").pack(side="left", padx=(0, 8))
        
        # Добавляем кнопку подсчета времени атаки в блок проверки атаки
        ttk.Button(button_frame1, text="⏱️ Подсчитать время атаки", 
                  command=self.calculate_attack_time, 
                  style="Rounded.TButton").pack(side="left")
        
        # Фрейм для проверки ID
        id_check_frame = self.create_rounded_frame(left_frame)
        id_check_frame.pack(fill="x", padx=10, pady=8)
        
        tk.Label(id_check_frame, text="🆔 Проверка ID", 
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['surface']).pack(anchor="w", pady=(12, 8), padx=12)
        
        input_frame2 = tk.Frame(id_check_frame, bg=self.colors['surface'])
        input_frame2.pack(fill="x", padx=12, pady=8)
        
        tk.Label(input_frame2, text="Папка ID:", 
                font=("Segoe UI", 9),
                bg=self.colors['surface']).grid(row=0, column=0, sticky="w")
        
        self.id_check_entry = ttk.Entry(input_frame2, font=("Segoe UI", 9))
        self.id_check_entry.grid(row=0, column=1, sticky="ew", padx=8)
        
        ttk.Button(input_frame2, text="Обзор", 
                  command=lambda: self.browse_folder(self.id_check_entry),
                  style="Secondary.TButton").grid(row=0, column=2, padx=(5, 0))
        
        input_frame2.columnconfigure(1, weight=1)
        
        button_frame2 = tk.Frame(id_check_frame, bg=self.colors['surface'])
        button_frame2.pack(fill="x", padx=12, pady=8)
        
        ttk.Button(button_frame2, text="🔍 Проверить ID", 
                  command=self.check_id, 
                  style="Rounded.TButton").pack(side="left", padx=(0, 8))
        
        # Добавляем кнопку подсчета времени ID в блок проверки ID
        ttk.Button(button_frame2, text="⏱️ Подсчитать время ID", 
                  command=self.calculate_id_time, 
                  style="Rounded.TButton").pack(side="left")
        
        # Фрейм для общей проверки
        global_check_frame = self.create_rounded_frame(left_frame)
        global_check_frame.pack(fill="x", padx=10, pady=8)
        
        tk.Label(global_check_frame, text="🌐 Общая проверка", 
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['surface']).pack(anchor="w", pady=(12, 8), padx=12)
        
        input_frame3 = tk.Frame(global_check_frame, bg=self.colors['surface'])
        input_frame3.pack(fill="x", padx=12, pady=8)
        
        tk.Label(input_frame3, text="Общая папка проекта:", 
                font=("Segoe UI", 9),
                bg=self.colors['surface']).grid(row=0, column=0, sticky="w")
        
        self.global_check_entry = ttk.Entry(input_frame3, font=("Segoe UI", 9))
        self.global_check_entry.grid(row=0, column=1, sticky="ew", padx=8)
        
        ttk.Button(input_frame3, text="Обзор", 
                  command=lambda: self.browse_folder(self.global_check_entry),
                  style="Secondary.TButton").grid(row=0, column=2, padx=(5, 0))
        
        input_frame3.columnconfigure(1, weight=1)
        
        button_frame3 = tk.Frame(global_check_frame, bg=self.colors['surface'])
        button_frame3.pack(fill="x", padx=12, pady=8)
        
        ttk.Button(button_frame3, text="🔍 Выполнить общую проверку", 
                  command=self.check_global, 
                  style="Rounded.TButton").pack(side="left", padx=(0, 8))
        
        # Добавляем кнопку подсчета времени проекта в блок общей проверки
        ttk.Button(button_frame3, text="⏱️ Подсчитать время проекта", 
                  command=self.calculate_project_time, 
                  style="Rounded.TButton").pack(side="left")
        
        # Фрейм для выгрузки отчёта
        report_frame = self.create_rounded_frame(left_frame)
        report_frame.pack(fill="x", padx=10, pady=8)
        
        tk.Label(report_frame, text="📊 Выгрузка отчёта", 
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['surface']).pack(anchor="w", pady=(12, 8), padx=12)
        
        # Кнопка для выгрузки отчёта
        ttk.Button(report_frame, text="📊 Выгрузить отчёт в Excel", 
                  command=self.export_shooting_report, 
                  style="Success.TButton").pack(fill="x", padx=12, pady=8)
        
        # Настройка правой части - логов проверки
        check_log_header = tk.Frame(right_frame, bg=self.colors['surface'])
        check_log_header.pack(fill="x", padx=12, pady=(12, 8))
        
        tk.Label(check_log_header, text="📋 Логи проверки", 
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['surface']).pack(side="left")
        
        # Кнопка очистки логов в заголовке
        ttk.Button(check_log_header, text="🧹 Очистить логи", 
                  command=self.clear_check_logs, style="Secondary.TButton").pack(side="right")
        
        # Контейнер для логов проверки
        check_log_container = tk.Frame(right_frame, bg=self.colors['surface'])
        check_log_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        
        self.check_log_text = scrolledtext.ScrolledText(check_log_container, height=20, font=("Consolas", 8),
                                                       bg='#1e293b', fg='#e2e8f0', 
                                                       insertbackground='white',
                                                       relief='flat',
                                                       padx=8, pady=8)
        self.check_log_text.pack(fill="both", expand=True)
        
        # Настройка тегов для цветного текста
        self.check_log_text.tag_config("SUCCESS", foreground="#10b981")
        self.check_log_text.tag_config("WARNING", foreground="#f59e0b")
        self.check_log_text.tag_config("ERROR", foreground="#ef4444")
        self.check_log_text.tag_config("INFO", foreground="#e2e8f0")
        self.check_log_text.tag_config("CRITICAL", foreground="#ff0000", background="#330000")
        self.check_log_text.tag_config("HEADER", foreground="#93c5fd", font=("Consolas", 8, "bold"))
        self.check_log_text.tag_config("SECTION", foreground="#cbd5e1", font=("Consolas", 8, "bold"))
        self.check_log_text.tag_config("DETAIL", foreground="#94a3b8")
    
    def setup_settings_tab(self, parent):
        edit_frame = self.create_rounded_frame(parent)
        edit_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Label(edit_frame, text="⚙️ Управление настройками атак", 
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['surface']).pack(anchor="w", pady=(12, 15), padx=12)
        
        input_frame1 = tk.Frame(edit_frame, bg=self.colors['surface'])
        input_frame1.pack(fill="x", padx=12, pady=8)
        
        tk.Label(input_frame1, text="Атака:", 
                font=("Segoe UI", 9),
                bg=self.colors['surface']).grid(row=0, column=0, sticky="w")
        
        self.edit_attack_var = tk.StringVar()
        self.edit_attack_combo = ttk.Combobox(input_frame1, textvariable=self.edit_attack_var, 
                                            values=list(self.attack_ranges.keys()), 
                                            state="readonly", font=("Segoe UI", 9))
        self.edit_attack_combo.grid(row=0, column=1, sticky="ew", padx=8)
        self.edit_attack_combo.bind("<<ComboboxSelected>>", self.load_attack_data)
        
        input_frame1.columnconfigure(1, weight=1)
        
        input_frame2 = tk.Frame(edit_frame, bg=self.colors['surface'])
        input_frame2.pack(fill="x", padx=12, pady=8)
        
        tk.Label(input_frame2, text="Kozen 10 (начало-конец):", 
                font=("Segoe UI", 9),
                bg=self.colors['surface']).grid(row=0, column=0, sticky="w", pady=4)
        
        self.kozen10_entry = ttk.Entry(input_frame2, font=("Segoe UI", 9))
        self.kozen10_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=4)
        
        tk.Label(input_frame2, text="Kozen 12 (начало-конец):", 
                font=("Segoe UI", 9),
                bg=self.colors['surface']).grid(row=1, column=0, sticky="w", pady=4)
        
        self.kozen12_entry = ttk.Entry(input_frame2, font=("Segoe UI", 9))
        self.kozen12_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=4)
        
        input_frame2.columnconfigure(1, weight=1)
        
        button_frame = tk.Frame(edit_frame, bg=self.colors['surface'])
        button_frame.pack(fill="x", padx=12, pady=15)
        
        # Кнопки в две строки для маленьких экранов
        top_button_frame = tk.Frame(button_frame, bg=self.colors['surface'])
        top_button_frame.pack(fill="x", pady=2)
        
        ttk.Button(top_button_frame, text="💾 Сохранить", 
                  command=self.save_attack_data, style="Success.TButton").pack(side="left", padx=2)
        ttk.Button(top_button_frame, text="➕ Новая атака", 
                  command=self.new_attack, style="Rounded.TButton").pack(side="left", padx=2)
        
        bottom_button_frame = tk.Frame(button_frame, bg=self.colors['surface'])
        bottom_button_frame.pack(fill="x", pady=2)
        
        ttk.Button(bottom_button_frame, text="✏️ Переименовать", 
                  command=self.rename_attack, style="Secondary.TButton").pack(side="left", padx=2)
        ttk.Button(bottom_button_frame, text="🗑️ Удалить атаку", 
                  command=self.delete_attack, style="Secondary.TButton").pack(side="left", padx=2)
    
    def browse_source(self):
        folder = filedialog.askdirectory()
        if folder:
            self.source_entry.delete(0, tk.END)
            self.source_entry.insert(0, folder)
            self.update_source_info()
    
    def browse_dest(self):
        folder = filedialog.askdirectory()
        if folder:
            # Проверяем, содержит ли имя папки "id" (без учета регистра)
            folder_name = os.path.basename(folder).lower()
            if "id" not in folder_name:
                result = messagebox.askyesno(
                    "Подтверждение", 
                    "Для корректной выгрузки в данном поле надо выбрать ТОЛЬКО папку самого ID. Вы уверены, что хотите использовать именно эту папку?",
                    icon="warning"
                )
                if not result:
                    return
            
            self.dest_entry.delete(0, tk.END)
            self.dest_entry.insert(0, folder)
    
    def browse_folder(self, entry_widget):
        folder = filedialog.askdirectory()
        if folder:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, folder)
    
    def update_source_info(self, event=None):
        """Обновляет информацию о количестве папок в исходной папке"""
        source_folder = self.source_entry.get()
        if source_folder and os.path.exists(source_folder):
            try:
                folders = [f for f in os.listdir(source_folder) 
                          if os.path.isdir(os.path.join(source_folder, f))]
                count = len(folders)
                self.source_info_label.config(text=f"📁 Найдено папок: {count}")
                
                # Также обновляем информацию о диапазоне
                self.update_range_info()
            except Exception:
                self.source_info_label.config(text="❌ Ошибка доступа к папке")
        else:
            self.source_info_label.config(text="")
    
    def update_range_info(self, event=None):
        attack = self.attack_var.get()
        device = self.device_var.get()
        
        if device == "все":
            # Для режима "все" показываем отдельные диапазоны для каждого устройства
            range_info_parts = []
            total_folders = 0
            
            for device_name in ["kozen 10", "kozen 12"]:
                if attack in self.attack_ranges and device_name in self.attack_ranges[attack]:
                    start, end = self.attack_ranges[attack][device_name]
                    device_total = end - start + 1
                    total_folders += device_total
                    range_info_parts.append(f"{device_name}: {start}-{end} ({device_total} номеров)")
            
            if range_info_parts:
                range_text = " | ".join(range_info_parts)
                self.range_info.config(text=f"📊 Диапазоны по устройствам: {range_text} | Всего: {total_folders} номеров")
            else:
                self.range_info.config(text="❌ Нет данных о диапазонах")
        elif attack in self.attack_ranges and device in self.attack_ranges[attack]:
            start, end = self.attack_ranges[attack][device]
            total = end - start + 1
            self.range_info.config(text=f"📊 Диапазон: {start}-{end} (всего: {total} номеров)")
        else:
            self.range_info.config(text="❌ Выбранная комбинация недоступна")
    
    def get_image_shooting_date(self, image_path):
        """Получает дату съёмки из EXIF данных изображений"""
        try:
            # Для получения EXIF данных потребуется установка Pillow
            # pip install Pillow
            from PIL import Image
            from PIL.ExifTags import TAGS
            
            with Image.open(image_path) as img:
                exif_data = img._getexif()
                if exif_data:
                    for tag_id, value in exif_data.items():
                        tag = TAGS.get(tag_id, tag_id)
                        if tag == 'DateTimeOriginal':
                            # Формат: "2023:10:15 14:30:25"
                            return datetime.datetime.strptime(value, '%Y:%m:%d %H:%M:%S')
        except Exception:
            pass
        
        return None
    
    def find_bestshot_file(self, folder_path):
        """Находит файл BestShot в папке"""
        try:
            for file in os.listdir(folder_path):
                if "bestshot" in file.lower() and any(file.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png']):
                    return os.path.join(folder_path, file)
        except Exception:
            pass
        return None
    
    def find_image_files(self, folder_path):
        """Находит все файлы изображений в папке"""
        image_extensions = {'.jpg', '.jpeg', '.png', '.tiff', '.bmp', '.gif'}
        image_files = []
        
        try:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if any(file.lower().endswith(ext) for ext in image_extensions):
                        image_files.append(os.path.join(root, file))
        except Exception:
            pass
        
        return image_files
    
    def get_folder_shooting_date(self, folder_path):
        """Получает дату съёмки для папки на основе EXIF данных изображений"""
        # Сначала ищем BestShot
        bestshot_file = self.find_bestshot_file(folder_path)
        if bestshot_file:
            date = self.get_image_shooting_date(bestshot_file)
            if date:
                return date
        
        # Если BestShot не найден или нет EXIF данных, ищем в папках Captures и Focus
        subfolders_to_check = ['Captures', 'Focus']
        
        for subfolder in subfolders_to_check:
            subfolder_path = os.path.join(folder_path, subfolder)
            if os.path.exists(subfolder_path):
                image_files = self.find_image_files(subfolder_path)
                # Сортируем файлы по имени для последовательности
                image_files.sort()
                for image_file in image_files:
                    date = self.get_image_shooting_date(image_file)
                    if date:
                        return date
        
        # Если ничего не найдено, ищем любые изображения в папке
        image_files = self.find_image_files(folder_path)
        # Сортируем файлы по имени для последовательности
        image_files.sort()
        for image_file in image_files:
            date = self.get_image_shooting_date(image_file)
            if date:
                return date
        
        return None
    
    def get_folder_shooting_time(self, folder_path):
        """Получает время съёмки для папки на основе EXIF данных изображений"""
        # Сначала ищем BestShot
        bestshot_file = self.find_bestshot_file(folder_path)
        if bestshot_file:
            date = self.get_image_shooting_date(bestshot_file)
            if date:
                return date
        
        # Если BestShot не найден, ищем в папках Captures и Focus
        subfolders_to_check = ['Captures', 'Focus']
        
        for subfolder in subfolders_to_check:
            subfolder_path = os.path.join(folder_path, subfolder)
            if os.path.exists(subfolder_path):
                image_files = self.find_image_files(subfolder_path)
                # Сортируем файлы по имени для последовательности
                image_files.sort()
                for image_file in image_files:
                    date = self.get_image_shooting_date(image_file)
                    if date:
                        return date
        
        # Если ничего не найдено, ищем любые изображения в папке
        image_files = self.find_image_files(folder_path)
        # Сортируем файлы по имени для последовательности
        image_files.sort()
        for image_file in image_files:
            date = self.get_image_shooting_date(image_file)
            if date:
                return date
        
        return None
    
    def calculate_shooting_time_for_folders(self, folder_paths):
        """Вычисляет время съёмки на основе дат съёмки из EXIF данных для списка путей к папкам"""
        if not folder_paths:
            return "не удалось вычислить"
        
        try:
            # Получаем времена съёмки всех папок из EXIF
            shooting_times = []
            for folder_path in folder_paths:
                shooting_time = self.get_folder_shooting_time(folder_path)
                if shooting_time:
                    folder_name = os.path.basename(folder_path)
                    shooting_times.append((folder_name, shooting_time))
            
            if not shooting_times:
                return "не удалось вычислить"
            
            # Сортируем по времени съёмки
            shooting_times.sort(key=lambda x: x[1])
            
            # Группируем по дням
            days_dict = {}
            for folder, time_obj in shooting_times:
                day_key = time_obj.date()
                if day_key not in days_dict:
                    days_dict[day_key] = []
                days_dict[day_key].append((folder, time_obj))
            
            total_seconds = 0
            day_count = 0
            
            # Обрабатываем каждый день отдельно
            for day, day_times in days_dict.items():
                day_count += 1
                day_times.sort(key=lambda x: x[1])
                
                # Разбиваем на сессии внутри дня (группы с интервалом менее 2 часов)
                sessions = []
                current_session = [day_times[0]]
                
                for i in range(1, len(day_times)):
                    time_diff = (day_times[i][1] - day_times[i-1][1]).total_seconds()
                    if time_diff > 7200:  # 2 часа в секундах
                        sessions.append(current_session)
                        current_session = [day_times[i]]
                    else:
                        current_session.append(day_times[i])
                
                sessions.append(current_session)
                
                # Вычисляем время съёмки для дня
                day_seconds = 0
                for session in sessions:
                    if len(session) > 1:
                        first_time = session[0][1].timestamp()
                        last_time = session[-1][1].timestamp()
                        session_duration = last_time - first_time
                        day_seconds += session_duration
                    else:
                        # Для одиночных сессий используем минимальное время 30 секунд
                        day_seconds += 30
                
                total_seconds += day_seconds
                
                # Логируем информацию о дне
                first_dt = day_times[0][1]
                last_dt = day_times[-1][1]
                self.check_log(f"📅 День {day_count} ({first_dt.strftime('%Y-%m-%d')}): {len(day_times)} папок, время: {self.format_duration(day_seconds)}", "DETAIL")
                
                # Логируем сессии внутри дня
                for i, session in enumerate(sessions, 1):
                    if len(session) > 1:
                        first_session_time = session[0][1]
                        last_session_time = session[-1][1]
                        session_duration = last_session_time.timestamp() - first_session_time.timestamp()
                        self.check_log(f"  📊 Сессия {i}: {first_session_time.strftime('%H:%M:%S')} - {last_session_time.strftime('%H:%M:%S')} "
                               f"({len(session)} папок, время: {self.format_duration(session_duration)})", "DETAIL")
                    else:
                        self.check_log(f"  📊 Сессия {i}: 1 папка, время: 00:00:30", "DETAIL")
            
            if total_seconds == 0:
                # Минимальное время съемки - 30 секунд на папку
                return self.format_duration(len(folder_paths) * 30)
            
            return self.format_duration(total_seconds)
            
        except Exception as e:
            self.check_log(f"Ошибка вычисления времени съёмки: {str(e)}", "WARNING")
            return "не удалось вычислить"
    
    def format_duration(self, total_seconds):
        """Форматирует длительность в формат HH:MM:SS"""
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    def get_common_shooting_date(self, folder_paths):
        """Получает самую частую дату съёмки из списка папок на основе EXIF данных"""
        all_shooting_dates = []
        
        for folder_path in folder_paths:
            folder_date = self.get_folder_shooting_date(folder_path)
            if folder_date:
                all_shooting_dates.append(folder_date)
        
        if not all_shooting_dates:
            return "не удалось определить"
        
        # Считаем частоту дат
        date_counts = {}
        for date in all_shooting_dates:
            date_str = date.strftime("%Y-%m-%d")
            date_counts[date_str] = date_counts.get(date_str, 0) + 1
        
        # Находим дату с максимальным количеством упоминаний
        common_date = max(date_counts.items(), key=lambda x: x[1])[0]
        return common_date
    
    def parse_number_range(self, range_str):
        """Парсинг диапазона номеров с сохранением порядка ввода"""
        numbers = []
        parts = [part.strip() for part in range_str.split(',')]
        
        for part in parts:
            if not part:
                continue
                
            if '-' in part:
                range_parts = part.split('-')
                if len(range_parts) != 2:
                    return None
                
                try:
                    start = int(range_parts[0].strip())
                    end = int(range_parts[1].strip())
                    
                    if start <= end:
                        numbers.extend(range(start, end + 1))
                    else:
                        numbers.extend(range(start, end - 1, -1))
                except ValueError:
                    return None
            else:
                try:
                    numbers.append(int(part))
                except ValueError:
                    return None
        
        return numbers
    
    def natural_sort_key(self, s):
        """Ключ для естественной сортировки как в проводнике Windows"""
        return [int(text) if text.isdigit() else text.lower()
                for text in re.split('([0-9]+)', s)]
    
    def get_attack_expected_count(self, attack_name, device):
        """Получает ожидаемое количество папок для атаки и устройства"""
        if attack_name not in self.attack_ranges:
            return 0
        
        if device == "все":
            # Для режима "все" возвращаем сумму количеств для всех устройств
            total = 0
            for device_name in ["kozen 10", "kozen 12"]:
                if device_name in self.attack_ranges[attack_name]:
                    start, end = self.attack_ranges[attack_name][device_name]
                    total += (end - start + 1)
            return total
        else:
            if device in self.attack_ranges[attack_name]:
                start, end = self.attack_ranges[attack_name][device]
                return end - start + 1
            return 0

    def is_numeric_folder(self, folder_name):
        """Проверяет что имя папки числовое (1-4 цифры)"""
        return folder_name.isdigit() and 1 <= len(folder_name) <= 4

    def log(self, message, level="INFO"):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if level == "WARNING":
            icon = "⚠️"
            tag = "WARNING"
        elif level == "ERROR":
            icon = "❌"
            tag = "ERROR"
        elif level == "SUCCESS":
            icon = "✅"
            tag = "SUCCESS"
        elif level == "CRITICAL":
            icon = "🚫"
            tag = "CRITICAL"
        elif level == "HEADER":
            icon = "📋"
            tag = "HEADER"
        elif level == "DETAIL":
            icon = "  📄"
            tag = "DETAIL"
        else:
            icon = "ℹ️"
            tag = "INFO"
        
        formatted_message = f"[{timestamp}] {icon} {message}\n"
        
        self.log_text.insert(tk.END, formatted_message, tag)
        self.log_text.see(tk.END)
        self.root.update()
    
    def check_log(self, message, level="INFO", indent=0):
        """Логирование для вкладки проверки с поддержкой отступов"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        
        if level == "WARNING":
            icon = "⚠️"
            tag = "WARNING"
        elif level == "ERROR":
            icon = "❌"
            tag = "ERROR"
        elif level == "SUCCESS":
            icon = "✅"
            tag = "SUCCESS"
        elif level == "CRITICAL":
            icon = "🚫"
            tag = "CRITICAL"
        elif level == "HEADER":
            icon = "📋"
            tag = "HEADER"
        elif level == "SECTION":
            icon = "  📁"
            tag = "SECTION"
        elif level == "DETAIL":
            icon = "    📄"
            tag = "DETAIL"
        else:
            icon = "ℹ️"
            tag = "INFO"
        
        indent_str = "  " * indent
        formatted_message = f"{indent_str}{icon} {message}\n"
        
        self.check_log_text.insert(tk.END, formatted_message, tag)
        self.check_log_text.see(tk.END)
        self.root.update()
    
    def clear_logs(self):
        self.log_text.delete(1.0, tk.END)
        self.log("Логи очищены", "INFO")
    
    def clear_check_logs(self):
        self.check_log_text.delete(1.0, tk.END)
        self.check_log("Логи проверки очищены", "INFO")
    
    def check_folder_content(self, folder_path, log_errors=True, indent=0, check_names=False, log_to_main=False):
        """
        Проверка содержимого папки
        check_names: если True, проверяет что имена папок числовые (для проверки атак)
        log_to_main: если True, логирует в основной лог вместо лога проверки
        """
        try:
            items = os.listdir(folder_path)
            folders = [item for item in items if os.path.isdir(os.path.join(folder_path, item))]
            files = [item for item in items if os.path.isfile(os.path.join(folder_path, item))]
            
            errors = []
            warnings = []
            
            # Проверка количества папок
            if len(folders) != 3:
                errors.append(f"Найдено {len(folders)} папок вместо 3")
            
            # Проверка наличия BestShot файла
            bestshot_files = [f for f in files if "bestshot" in f.lower()]
            if not bestshot_files:
                errors.append("Не найден файл BestShot")
            elif len(bestshot_files) > 1:
                warnings.append(f"Найдено {len(bestshot_files)} файлов BestShot")
            
            # Проверка что папки не пустые
            for folder in folders:
                folder_full_path = os.path.join(folder_path, folder)
                try:
                    if not os.listdir(folder_full_path):
                        errors.append(f"Папка '{folder}' пустая")
                except PermissionError:
                    errors.append(f"Нет доступа к папке '{folder}'")
            
            # Проверка числовых имен (только при check_names=True)
            if check_names:
                non_numeric = [f for f in folders if not self.is_numeric_folder(f)]
                if non_numeric:
                    errors.append(f"Нечисловые имена папок: {', '.join(non_numeric)}")
            
            if log_errors:
                if log_to_main:
                    # Логирование в основной лог
                    if errors:
                        for error in errors:
                            self.log(f"Ошибка в папке {os.path.basename(folder_path)}: {error}", "ERROR")
                    if warnings:
                        for warning in warnings:
                            self.log(f"Предупреждение в папке {os.path.basename(folder_path)}: {warning}", "WARNING")
                    if not errors and not warnings:
                        self.log(f"Папка {os.path.basename(folder_path)}: содержимое в порядке", "SUCCESS")
                else:
                    # Логирование в лог проверки
                    if errors:
                        for error in errors:
                            self.check_log(f"Ошибка: {error}", "ERROR", indent)
                    if warnings:
                        for warning in warnings:
                            self.check_log(f"Предупреждение: {warning}", "WARNING", indent)
                    if not errors and not warnings:
                        self.check_log("Содержимое папки в порядке", "SUCCESS", indent)
            
            return len(errors) == 0
            
        except Exception as e:
            if log_errors:
                error_msg = f"Ошибка проверки папки: {str(e)}"
                if log_to_main:
                    self.log(f"Ошибка проверки папки {os.path.basename(folder_path)}: {str(e)}", "ERROR")
                else:
                    self.check_log(error_msg, "ERROR", indent)
            return False

    def execute_renaming(self):
        source_folder = self.source_entry.get()
        dest_folder = self.dest_entry.get()
        device = self.device_var.get()
        attack = self.attack_var.get()
        check_content = self.check_content_var.get()
        
        if not source_folder or not dest_folder:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите исходную папку и папку назначения")
            return
        
        if not os.path.exists(source_folder):
            messagebox.showerror("Ошибка", "Исходная папка не существует")
            return
        
        all_folders = [f for f in os.listdir(source_folder) 
                      if os.path.isdir(os.path.join(source_folder, f))]
        
        all_folders.sort(key=self.natural_sort_key)
        
        if not all_folders:
            messagebox.showwarning("Предупреждение", "В исходной папке не найдено папок для обработки")
            return
        
        # Время съемки считается ТОЛЬКО для обрабатываемых папок
        folder_paths = [os.path.join(source_folder, folder) for folder in all_folders]
        shooting_time = self.calculate_shooting_time_for_folders(folder_paths)
        
        if device != "все" and (attack not in self.attack_ranges or device not in self.attack_ranges[attack]):
            messagebox.showerror("Ошибка", f"Выбранная комбинация атаки {attack} и устройства {device} недоступна")
            return
        
        # ОПРЕДЕЛЯЕМ СКОЛЬКО ПАПОК БУДЕМ ОБРАБАТЫВАТЬ
        expected_count = self.get_attack_expected_count(attack, device)
        actual_count = len(all_folders)
        
        # Определяем сколько папок будем обрабатывать
        if expected_count > 0:
            if actual_count > expected_count:
                self.log(f"⚠️ Внимание: в исходной папке {actual_count} папок, но требуется только {expected_count}", "WARNING")
                self.log(f"ℹ️ Будет обработано только {expected_count} папок", "INFO")
                folders_to_process = all_folders[:expected_count]  # Берем только нужное количество
                processing_count = expected_count
            else:
                folders_to_process = all_folders
                processing_count = actual_count
        else:
            folders_to_process = all_folders
            processing_count = actual_count
        
        try:
            os.makedirs(dest_folder, exist_ok=True)
            attack_folder = os.path.join(dest_folder, attack)
            os.makedirs(attack_folder, exist_ok=True)
            
            self.log("=" * 70, "SUCCESS")
            self.log(f"🚀 Начало обработки...", "HEADER")
            self.log(f"📊 Найдено папок для обработки: {processing_count}", "INFO")
            if expected_count > 0:
                self.log(f"📋 Ожидаемое количество для атаки: {expected_count}", "INFO")
            
            # ПРЕДВАРИТЕЛЬНАЯ ПРОВЕРКА СОДЕРЖИМОГО
            if check_content:
                self.log("🔍 Начинается предварительная проверка содержимого...", "INFO")
                content_errors = False
                error_details = []
                
                for i, folder in enumerate(folders_to_process, 1):
                    old_path = os.path.join(source_folder, folder)
                    self.log(f"🔍 Проверка {i}/{len(folders_to_process)}: {folder}", "DETAIL")
                    
                    # Подробная проверка с выводом ошибок в ОСНОВНОЙ лог
                    try:
                        if not self.check_folder_content(old_path, log_errors=True, indent=1, check_names=False, log_to_main=True):
                            content_errors = True
                            error_details.append(folder)
                            self.log(f"❌ Обнаружены ошибки в папке: {folder}", "ERROR")
                        else:
                            self.log(f"✅ Папка {folder} проверена успешно", "SUCCESS")
                    except Exception as e:
                        content_errors = True
                        error_details.append(folder)
                        self.log(f"❌ Ошибка при проверке папке {folder}: {str(e)}", "ERROR")
                
                if content_errors:
                    self.log("🚫 ОБНАРУЖЕНЫ ОШИБКИ! Переименование отменено.", "ERROR")
                    self.log(f"📂 Папки с ошибками: {', '.join(error_details)}", "ERROR")
                    messagebox.showerror("Ошибка", 
                                        "Обнаружены ошибки в содержимом папок! "
                                        "Переименование отменено. Проверьте логи для деталей.")
                    return
                else:
                    self.log("✅ Все папки проверены успешно!", "SUCCESS")
            
            processed_count = 0
            
            if device == "все":
                # Создаем папки для устройств если их нет
                devices_in_attack = []
                for device_name in ["kozen 10", "kozen 12"]:
                    if attack in self.attack_ranges and device_name in self.attack_ranges[attack]:
                        devices_in_attack.append(device_name)
                        device_folder = os.path.join(attack_folder, device_name)
                        os.makedirs(device_folder, exist_ok=True)
                        self.log(f"📁 Создана папка устройства: {device_name}", "INFO")
                
                if len(devices_in_attack) == 0:
                    messagebox.showerror("Ошибка", f"Для атаки {attack} не заданы диапазоны")
                    return
                
                # НОВАЯ ЛОГИКА: если ровно 2 папки, то первая - kozen 10, вторая - kozen 12
                if len(folders_to_process) == 2 and len(devices_in_attack) >= 2:
                    self.log("🎯 Обнаружено 2 папки: первая для Kozen 10, вторая для Kozen 12", "INFO")
                    
                    # Обработка первой папки для Kozen 10
                    if "kozen 10" in devices_in_attack:
                        device_name = "kozen 10"
                        start_num, end_num = self.attack_ranges[attack][device_name]
                        device_folder = os.path.join(attack_folder, device_name)
                        
                        # Получаем все вложенные папки из первой папки
                        first_source_folder = os.path.join(source_folder, folders_to_process[0])
                        first_subfolders = [f for f in os.listdir(first_source_folder) 
                                          if os.path.isdir(os.path.join(first_source_folder, f))]
                        first_subfolders.sort(key=self.natural_sort_key)
                        
                        self.log(f"📁 Обработка папки {folders_to_process[0]} для {device_name}: {len(first_subfolders)} вложенных папок", "INFO")
                        
                        available_numbers = list(range(start_num, end_num + 1))
                        actual_processing = min(len(first_subfolders), len(available_numbers))
                        
                        if actual_processing < len(first_subfolders):
                            self.log(f"⚠️ Для {device_name} доступно только {len(available_numbers)} номеров, обрабатываем {actual_processing} папок", "WARNING")
                        
                        current_number = start_num
                        
                        for i in range(actual_processing):
                            subfolder = first_subfolders[i]
                            old_path = os.path.join(first_source_folder, subfolder)
                            new_name = str(current_number)
                            new_path = os.path.join(device_folder, new_name)
                            
                            if os.path.exists(new_path):
                                shutil.rmtree(new_path)
                                self.log(f"Удалена существующая папка: {device_name}/{new_name}", "WARNING")
                            
                            shutil.copytree(old_path, new_path)
                            self.log(f"Обработано: {folders_to_process[0]}/{subfolder} → {device_name}/{new_name}", "SUCCESS")
                            processed_count += 1
                            current_number += 1
                    
                    # Обработка второй папки для Kozen 12
                    if "kozen 12" in devices_in_attack:
                        device_name = "kozen 12"
                        start_num, end_num = self.attack_ranges[attack][device_name]
                        device_folder = os.path.join(attack_folder, device_name)
                        
                        # Получаем все вложенные папки из второй папки
                        second_source_folder = os.path.join(source_folder, folders_to_process[1])
                        second_subfolders = [f for f in os.listdir(second_source_folder) 
                                           if os.path.isdir(os.path.join(second_source_folder, f))]
                        second_subfolders.sort(key=self.natural_sort_key)
                        
                        self.log(f"📁 Обработка папки {folders_to_process[1]} для {device_name}: {len(second_subfolders)} вложенных папок", "INFO")
                        
                        available_numbers = list(range(start_num, end_num + 1))
                        actual_processing = min(len(second_subfolders), len(available_numbers))
                        
                        if actual_processing < len(second_subfolders):
                            self.log(f"⚠️ Для {device_name} доступно только {len(available_numbers)} номеров, обрабатываем {actual_processing} папок", "WARNING")
                        
                        current_number = start_num
                        
                        for i in range(actual_processing):
                            subfolder = second_subfolders[i]
                            old_path = os.path.join(second_source_folder, subfolder)
                            new_name = str(current_number)
                            new_path = os.path.join(device_folder, new_name)
                            
                            if os.path.exists(new_path):
                                shutil.rmtree(new_path)
                                self.log(f"Удалена существующая папка: {device_name}/{new_name}", "WARNING")
                            
                            shutil.copytree(old_path, new_path)
                            self.log(f"Обработано: {folders_to_process[1]}/{subfolder} → {device_name}/{new_name}", "SUCCESS")
                            processed_count += 1
                            current_number += 1
                        
                elif len(devices_in_attack) == 1:
                    # Если только одно устройство - все папки в него
                    device_name = devices_in_attack[0]
                    start_num, end_num = self.attack_ranges[attack][device_name]
                    available_numbers = list(range(start_num, end_num + 1))
                    
                    # Обрабатываем только доступное количество папок
                    actual_processing = min(len(folders_to_process), len(available_numbers))
                    if actual_processing < len(folders_to_process):
                        self.log(f"⚠️ Доступно только {len(available_numbers)} номеров, обрабатываем {actual_processing} папок", "WARNING")
                    
                    current_number = start_num
                    device_folder = os.path.join(attack_folder, device_name)
                    
                    for i in range(actual_processing):
                        folder = folders_to_process[i]
                        old_path = os.path.join(source_folder, folder)
                        new_name = str(current_number)
                        new_path = os.path.join(device_folder, new_name)
                        
                        if os.path.exists(new_path):
                            shutil.rmtree(new_path)
                            self.log(f"Удалена существующая папка: {new_name}", "WARNING")
                        
                        shutil.copytree(old_path, new_path)
                        self.log(f"Обработано: {folder} → {device_name}/{new_name}", "SUCCESS")
                        processed_count += 1
                        current_number += 1
                else:
                    # Если два устройства и не 2 папки - распределяем поровну
                    device1, device2 = devices_in_attack
                    start1, end1 = self.attack_ranges[attack][device1]
                    start2, end2 = self.attack_ranges[attack][device2]
                    
                    available_numbers1 = list(range(start1, end1 + 1))
                    available_numbers2 = list(range(start2, end2 + 1))
                    
                    # РАСПРЕДЕЛЯЕМ ПАПКИ ПОРОВНУ МЕЖДУ УСТРОЙСТВАМИ
                    half = len(folders_to_process) // 2
                    first_half = folders_to_process[:half]
                    second_half = folders_to_process[half:half * 2]  # Берем только нужное количество
                    
                    # Обработка первой половины для device1
                    actual_first_half = min(len(first_half), len(available_numbers1))
                    if actual_first_half < len(first_half):
                        self.log(f"⚠️ Для {device1} доступно только {len(available_numbers1)} номеров, обрабатываем {actual_first_half} папок", "WARNING")
                    
                    current_number = start1
                    device1_folder = os.path.join(attack_folder, device1)
                    
                    for i in range(actual_first_half):
                        folder = first_half[i]
                        old_path = os.path.join(source_folder, folder)
                        new_name = str(current_number)
                        new_path = os.path.join(device1_folder, new_name)
                        
                        if os.path.exists(new_path):
                            shutil.rmtree(new_path)
                            self.log(f"Удалена существующая папка: {device1}/{new_name}", "WARNING")
                        
                        shutil.copytree(old_path, new_path)
                        self.log(f"Обработано: {folder} → {device1}/{new_name}", "SUCCESS")
                        processed_count += 1
                        current_number += 1
                    
                    # Обработка второй половины для device2
                    actual_second_half = min(len(second_half), len(available_numbers2))
                    if actual_second_half < len(second_half):
                        self.log(f"⚠️ Для {device2} доступно только {len(available_numbers2)} номеров, обрабатываем {actual_second_half} папок", "WARNING")
                    
                    current_number = start2
                    device2_folder = os.path.join(attack_folder, device2)
                    
                    for i in range(actual_second_half):
                        folder = second_half[i]
                        old_path = os.path.join(source_folder, folder)
                        new_name = str(current_number)
                        new_path = os.path.join(device2_folder, new_name)
                        
                        if os.path.exists(new_path):
                            shutil.rmtree(new_path)
                            self.log(f"Удалена существующая папка: {device2}/{new_name}", "WARNING")
                        
                        shutil.copytree(old_path, new_path)
                        self.log(f"Обработано: {folder} → {device2}/{new_name}", "SUCCESS")
                        processed_count += 1
                        current_number += 1
            else:
                # Обработка для конкретного устройства
                device_folder = os.path.join(attack_folder, device)
                os.makedirs(device_folder, exist_ok=True)
                
                start_num, end_num = self.attack_ranges[attack][device]
                available_numbers = list(range(start_num, end_num + 1))
                
                # Обрабатываем только доступное количество папок
                actual_processing = min(len(folders_to_process), len(available_numbers))
                if actual_processing < len(folders_to_process):
                    self.log(f"⚠️ Доступно только {len(available_numbers)} номеров, обрабатываем {actual_processing} папок", "WARNING")
                
                current_number = start_num
                
                for i in range(actual_processing):
                    folder = folders_to_process[i]
                    old_path = os.path.join(source_folder, folder)
                    new_name = str(current_number)
                    new_path = os.path.join(device_folder, new_name)
                    
                    if os.path.exists(new_path):
                        shutil.rmtree(new_path)
                        self.log(f"Удалена существующая папка: {new_name}", "WARNING")
                    
                    shutil.copytree(old_path, new_path)
                    self.log(f"Обработано: {folder} → {new_name}", "SUCCESS")
                    processed_count += 1
                    current_number += 1
            
            self.log("=" * 70, "SUCCESS")
            self.log(f"✅ Обработка завершена успешно! Обработано: {processed_count} папок", "SUCCESS")
            self.log(f"⏱️ Общее время съёмки: {shooting_time}", "INFO")
            
            if len(all_folders) > processing_count:
                self.log(f"📝 Осталось необработанных папок: {len(all_folders) - processing_count}", "INFO")
            
            messagebox.showinfo("Успех", 
                               f"Обработка завершена!\n\n"
                               f"✅ Успешно обработано: {processed_count} папок\n"
                               f"⏱️ Время съёмки: {shooting_time}")
            
        except Exception as e:
            self.log(f"Ошибка: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
    
    def execute_replacement(self):
        source_folder = self.source_entry.get()
        dest_folder = self.dest_entry.get()
        device = self.device_var.get()
        attack = self.attack_var.get()
        replace_numbers_str = self.replace_entry.get()
        check_content = self.check_content_var.get()
        
        if not replace_numbers_str:
            messagebox.showerror("Ошибка", "Введите номера папок для замены")
            return
        
        replace_numbers = self.parse_number_range(replace_numbers_str)
        if replace_numbers is None:
            messagebox.showerror("Ошибка", "Неверный формат номеров. Используйте: 522,530-532,528")
            return
        
        source_folders = [f for f in os.listdir(source_folder) 
                        if os.path.isdir(os.path.join(source_folder, f))]
        
        source_folders.sort(key=self.natural_sort_key)
        
        if len(source_folders) != len(replace_numbers):
            messagebox.showerror("Ошибка", 
                f"Количество папок в исходной папке ({len(source_folders)}) "
                f"не соответствует количеству номеров для замены ({len(replace_numbers)})")
            return
        
        # Время съемки считается ТОЛЬКО для обрабатываемых папок
        folder_paths = [os.path.join(source_folder, folder) for folder in source_folders]
        shooting_time = self.calculate_shooting_time_for_folders(folder_paths)
        
        if device != "все" and (attack not in self.attack_ranges or device not in self.attack_ranges[attack]):
            messagebox.showerror("Ошибка", f"Выбранная комбинация атаки {attack} и устройства {device} недоступна")
            return
        
        try:
            attack_folder = os.path.join(dest_folder, attack)
            
            if device == "все":
                # Определяем устройства в атаке
                devices_in_attack = []
                for device_name in ["kozen 10", "kozen 12"]:
                    if attack in self.attack_ranges and device_name in self.attack_ranges[attack]:
                        devices_in_attack.append(device_name)
                
                if len(devices_in_attack) == 0:
                    messagebox.showerror("Ошибка", f"Для атаки {attack} не заданы диапазоны")
                    return
                
                # Проверяем что все номера входят в соответствующие диапазоны
                for i, num in enumerate(replace_numbers):
                    found_device = None
                    for device_name in devices_in_attack:
                        start_num, end_num = self.attack_ranges[attack][device_name]
                        if start_num <= num <= end_num:
                            found_device = device_name
                            break
                    
                    # Для атак 10-15 пропускаем проверку диапазонов
                    if not found_device and attack not in ["10 Indoors", "11 Indoors. With attributes", "12 Indoors. Backlight", 
                                        "13 Indoors. Insufficient lighting", "14 Indoors. Behind transparent glass", "15 Outside"]:
                        messagebox.showerror("Ошибка", f"Номер {num} не входит ни в один диапазон атаки {attack}")
                        return
            else:
                # Для обычных атак проверяем диапазон
                if attack not in ["10 Indoors", "11 Indoors. With attributes", "12 Indoors. Backlight", 
                                "13 Indoors. Insufficient lighting", "14 Indoors. Behind transparent glass", "15 Outside"]:
                    start_num, end_num = self.attack_ranges[attack][device]
                    for num in replace_numbers:
                        if num < start_num or num > end_num:
                            messagebox.showerror("Ошибка", f"Номер {num} вне диапазона {start_num}-{end_num}")
                            return
            
            if not os.path.exists(attack_folder):
                messagebox.showerror("Ошибка", f"Папка назначения {attack_folder} не существует")
                return
            
            self.log("=" * 70, "SUCCESS")
            self.log(f"🔄 Начало замены папок...", "HEADER")
            self.log(f"🔢 Заменяемые номера: {replace_numbers}", "INFO")
            
            # ПРЕДВАРИТЕЛЬНАЯ ПРОВЕРКА СОДЕРЖИМОГО
            if check_content:
                self.log("🔍 Начинается предварительная проверка содержимого...", "INFO")
                content_errors = False
                error_details = []
                
                for i, folder in enumerate(source_folders, 1):
                    old_path = os.path.join(source_folder, folder)
                    self.log(f"🔍 Проверка {i}/{len(source_folders)}: {folder}", "DETAIL")
                    
                    try:
                        if not self.check_folder_content(old_path, log_errors=True, indent=1, check_names=False, log_to_main=True):
                            content_errors = True
                            error_details.append(folder)
                            self.log(f"❌ Обнаружены ошибки в папке: {folder}", "ERROR")
                        else:
                            self.log(f"✅ Папка {folder} проверена успешно", "SUCCESS")
                    except Exception as e:
                        content_errors = True
                        error_details.append(folder)
                        self.log(f"❌ Ошибка при проверке папки {folder}: {str(e)}", "ERROR")
                
                if content_errors:
                    self.log("🚫 ОБНАРУЖЕНЫ ОШИБКИ! Замена отменена.", "ERROR")
                    self.log(f"📂 Папки с ошибками: {', '.join(error_details)}", "ERROR")
                    messagebox.showerror("Ошибка", 
                                        "Обнаружены ошибки в содержимом папок! "
                                        "Замена отменена. Проверьте логи для деталей.")
                    return
                else:
                    self.log("✅ Все папки проверены успешно!", "SUCCESS")
            
            replaced_count = 0
            
            if device == "все":
                for i, folder in enumerate(source_folders):
                    old_path = os.path.join(source_folder, folder)
                    target_number = replace_numbers[i]
                    
                    # Определяем устройство для этого номера
                    found_device = None
                    for device_name in devices_in_attack:
                        start_num, end_num = self.attack_ranges[attack][device_name]
                        if start_num <= target_number <= end_num:
                            found_device = device_name
                            break
                    
                    # Если устройство не найдено (для атак 10-15), используем первое доступное
                    if not found_device and devices_in_attack:
                        found_device = devices_in_attack[0]
                    
                    if not found_device:
                        self.log(f"❌ Не удалось определить устройство для номера {target_number}", "ERROR")
                        continue
                    
                    device_folder = os.path.join(attack_folder, found_device)
                    new_name = str(target_number)
                    new_path = os.path.join(device_folder, new_name)
                    
                    if os.path.exists(new_path):
                        shutil.rmtree(new_path)
                    
                    shutil.copytree(old_path, new_path)
                    self.log(f"Заменено: {folder} → {found_device}/{new_name}", "SUCCESS")
                    replaced_count += 1
            else:
                device_folder = os.path.join(attack_folder, device)
                
                for i, folder in enumerate(source_folders):
                    old_path = os.path.join(source_folder, folder)
                    target_number = replace_numbers[i]
                    new_name = str(target_number)
                    new_path = os.path.join(device_folder, new_name)
                    
                    if os.path.exists(new_path):
                        shutil.rmtree(new_path)
                    
                    shutil.copytree(old_path, new_path)
                    self.log(f"Заменено: {folder} → {new_name}", "SUCCESS")
                    replaced_count += 1
            
            self.log("=" * 70, "SUCCESS")
            self.log(f"✅ Замена завершена успешно! Заменено: {replaced_count} папок", "SUCCESS")
            self.log(f"⏱️ Общее время съёмки: {shooting_time}", "INFO")
            
            messagebox.showinfo("Успех", 
                               f"Замена завершена!\n\n"
                               f"✅ Заменено папок: {replaced_count}\n"
                               f"⏱️ Время съёмки: {shooting_time}")
            
        except Exception as e:
            self.log(f"Ошибка при замене: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

    def check_attack(self):
        """Проверка отдельной атаки"""
        attack_folder = self.attack_check_entry.get()
        
        if not attack_folder:
            messagebox.showerror("Ошибка", "Выберите папку атаки")
            return
        
        if not os.path.exists(attack_folder):
            messagebox.showerror("Ошибка", "Папка атаки не существует")
            return
        
        self.check_log("=" * 60, "HEADER")
        self.check_log(f"🔍 ПРОВЕРКА АТАКИ: {os.path.basename(attack_folder)}", "HEADER")
        self.check_log("=" * 60, "HEADER")
        
        try:
            attack_name = os.path.basename(attack_folder)
            
            if attack_name not in self.attack_ranges:
                self.check_log(f"❌ ОШИБКА: Папка не является известной атакой", "ERROR")
                self.check_log(f"📝 Название папки: {attack_name}", "INFO")
                self.check_log(f"📋 Доступные атаки: {', '.join(self.attack_ranges.keys())}", "INFO")
                return
            
            attack_type = attack_name
            structure_info = self.check_attack_structure(attack_folder, attack_type)
            
            self.check_log(f"📁 Структура: {structure_info['structure_type']}", "INFO")
            self.check_log(f"📊 Ожидаемое количество папок: {structure_info['expected_total']}", "INFO")
            
            total_errors = 0
            total_folders = 0
            total_checked = 0
            
            if structure_info['has_kozen10'] or structure_info['has_kozen12']:
                for device in ["kozen 10", "kozen 12"]:
                    if structure_info[f'has_{device.replace(" ", "")}']:
                        device_folder = os.path.join(attack_folder, device)
                        
                        self.check_log(f"", "INFO")
                        self.check_log(f"📱 Устройство: {device}", "SECTION")
                        
                        expected_count = 0
                        if attack_type in self.attack_ranges and device in self.attack_ranges[attack_type]:
                            start, end = self.attack_ranges[attack_type][device]
                            expected_count = end - start + 1
                        
                        try:
                            all_items = os.listdir(device_folder)
                            # ПРИ ПРОВЕРКЕ АТАКИ проверяем числовые имена (1-4 цифры)
                            folders = [f for f in all_items 
                                      if os.path.isdir(os.path.join(device_folder, f)) and self.is_numeric_folder(f)]
                            
                            other_items = [item for item in all_items if item not in folders]
                            if other_items:
                                self.check_log(f"⚠️ Посторонние элементы: {', '.join(other_items)}", "WARNING", 1)
                            
                            actual_count = len(folders)
                            
                            self.check_log(f"📈 Ожидалось: {expected_count}", "INFO", 1)
                            self.check_log(f"📈 Найдено: {actual_count}", 
                                         "SUCCESS" if expected_count == actual_count else "ERROR", 1)
                            
                            if expected_count > 0 and actual_count != expected_count:
                                self.check_log(f"❌ НЕСООТВЕТСТВИЕ КОЛИЧЕСТВА!", "ERROR", 1)
                                total_errors += 1
                            
                            # ПРОВЕРКА СОДЕРЖИМОГО КАЖДОЙ ПАПКИ
                            self.check_log(f"🔍 Проверка содержимого папок:", "SECTION", 1)
                            folder_errors = 0
                            for folder in folders:
                                folder_path = os.path.join(device_folder, folder)
                                self.check_log(f"📂 Папка {folder}:", "DETAIL", 2)
                                # check_names=False - внутренние папки могут называться как угодно
                                if not self.check_folder_content(folder_path, log_errors=True, indent=3, check_names=False):
                                    folder_errors += 1
                                total_checked += 1
                            
                            total_errors += folder_errors
                            total_folders += actual_count
                            
                            if folder_errors == 0:
                                self.check_log(f"✅ Все папки устройства проверены успешно", "SUCCESS", 1)
                            else:
                                self.check_log(f"❌ Ошибок в папках: {folder_errors}", "ERROR", 1)
                        except Exception as e:
                            self.check_log(f"❌ Ошибка доступа к папке устройства: {str(e)}", "ERROR", 1)
                            total_errors += 1
            else:
                self.check_log(f"", "INFO")
                self.check_log(f"📁 Папки в корне атаки", "SECTION")
                
                try:
                    all_items = os.listdir(attack_folder)
                    # ПРИ ПРОВЕРКЕ АТАКИ проверяем числовые имена (1-4 цифры)
                    folders = [f for f in all_items 
                              if os.path.isdir(os.path.join(attack_folder, f)) and self.is_numeric_folder(f)]
                    
                    other_items = [item for item in all_items if item not in folders]
                    if other_items:
                        self.check_log(f"⚠️ Посторонние элементы: {', '.join(other_items)}", "WARNING", 1)
                    
                    actual_count = len(folders)
                    
                    self.check_log(f"📈 Ожидалось: {structure_info['expected_total']}", "INFO", 1)
                    self.check_log(f"📈 Найдено: {actual_count}", 
                                 "SUCCESS" if structure_info['expected_total'] == actual_count else "ERROR", 1)
                    
                    if structure_info['expected_total'] > 0 and actual_count != structure_info['expected_total']:
                        self.check_log(f"❌ НЕСООТВЕТСТВИЕ КОЛИЧЕСТВА!", "ERROR", 1)
                        total_errors += 1
                    
                    # ПРОВЕРКА СОДЕРЖИМОГО КАЖДОЙ ПАПКИ
                    self.check_log(f"🔍 Проверка содержимого папок:", "SECTION", 1)
                    folder_errors = 0
                    for folder in folders:
                        folder_path = os.path.join(attack_folder, folder)
                        self.check_log(f"📂 Папка {folder}:", "DETAIL", 2)
                        # check_names=False - внутренние папки могут называться как угодно
                        if not self.check_folder_content(folder_path, log_errors=True, indent=3, check_names=False):
                            folder_errors += 1
                        total_checked += 1
                    
                    total_errors += folder_errors
                    total_folders += actual_count
                    
                    if folder_errors == 0:
                        self.check_log(f"✅ Все папки проверены успешно", "SUCCESS", 1)
                    else:
                        self.check_log(f"❌ Ошибок в папках: {folder_errors}", "ERROR", 1)
                except Exception as e:
                    self.check_log(f"❌ Ошибка доступа к папке атаки: {str(e)}", "ERROR", 1)
                    total_errors += 1
            
            self.check_log("", "INFO")
            self.check_log("=" * 60, "HEADER")
            if total_errors == 0:
                self.check_log(f"✅ ПРОВЕРКА ЗАВЕРШЕНА УСПЕШНО!", "SUCCESS")
                self.check_log(f"📊 Проверено папок: {total_checked}", "SUCCESS")
                messagebox.showinfo("Проверка завершена", "Атака проверена успешно! Ошибок не обнаружено.")
            else:
                self.check_log(f"❌ ПРОВЕРКА ЗАВЕРШЕНА С ОШИБКАМИ", "ERROR")
                self.check_log(f"📊 Обнаружено ошибок: {total_errors}", "ERROR")
                messagebox.showwarning("Проверка завершена", f"Обнаружены ошибки: {total_errors}")
                
        except Exception as e:
            self.check_log(f"❌ Ошибка при проверке атаки: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка при проверке: {str(e)}")

    def check_attack_structure(self, attack_folder, attack_type):
        """Проверяет структуру папки атаки и возвращает информацию о ней"""
        try:
            items = os.listdir(attack_folder)
            
            has_kozen10 = "kozen 10" in items and os.path.isdir(os.path.join(attack_folder, "kozen 10"))
            has_kozen12 = "kozen 12" in items and os.path.isdir(os.path.join(attack_folder, "kozen 12"))
            
            structure_type = ""
            expected_total = self.get_attack_expected_count(attack_type, "все")
            
            if has_kozen10 or has_kozen12:
                structure_type = "раздельная (с устройствами)"
            else:
                structure_type = "плоская (все папки в корне)"
            
            return {
                "has_kozen10": has_kozen10,
                "has_kozen12": has_kozen12,
                "structure_type": structure_type,
                "expected_total": expected_total
            }
        except Exception as e:
            return {
                "has_kozen10": False,
                "has_kozen12": False,
                "structure_type": "ошибка доступа",
                "expected_total": 0
            }

    def check_id(self):
        """Проверка ID с проверкой содержимого папок"""
        id_folder = self.id_check_entry.get()
        
        if not id_folder:
            messagebox.showerror("Ошибка", "Выберите папку ID")
            return
        
        if not os.path.exists(id_folder):
            messagebox.showerror("Ошибка", "Папка ID не существует")
            return
        
        self.check_log("=" * 60, "HEADER")
        self.check_log(f"🆔 ПРОВЕРКА ID: {os.path.basename(id_folder)}", "HEADER")
        self.check_log("=" * 60, "HEADER")
        
        try:
            attack_folders = []
            for item in os.listdir(id_folder):
                item_path = os.path.join(id_folder, item)
                if os.path.isdir(item_path) and item in self.attack_ranges:
                    attack_folders.append((item, item_path))
            
            if not attack_folders:
                self.check_log(f"❌ ОШИБКА: В папке ID не найдено папок атак", "ERROR")
                self.check_log(f"📋 Доступные атаки: {', '.join(self.attack_ranges.keys())}", "INFO")
                return
            
            total_errors = 0
            total_attacks = len(attack_folders)
            total_content_errors = 0
            
            self.check_log(f"📊 Найдено атак: {total_attacks}", "INFO")
            self.check_log("", "INFO")
            
            for attack_name, attack_folder in attack_folders:
                self.check_log(f"🎯 Атака: {attack_name}", "SECTION")
                
                try:
                    structure_info = self.check_attack_structure(attack_folder, attack_name)
                    
                    self.check_log(f"📁 Структура: {structure_info['structure_type']}", "INFO", 1)
                    self.check_log(f"📊 Ожидаемое количество: {structure_info['expected_total']}", "INFO", 1)
                    
                    attack_errors = 0
                    content_errors = 0
                    actual_total = 0
                    
                    if structure_info['has_kozen10'] or structure_info['has_kozen12']:
                        for device in ["kozen 10", "kozen 12"]:
                            if structure_info[f'has_{device.replace(" ", "")}']:
                                device_folder = os.path.join(attack_folder, device)
                                
                                if not os.path.exists(device_folder):
                                    self.check_log(f"⚠️ Папка устройства {device} не существует", "WARNING", 2)
                                    continue
                                
                                expected_count = 0
                                if attack_name in self.attack_ranges and device in self.attack_ranges[attack_name]:
                                    start, end = self.attack_ranges[attack_name][device]
                                    expected_count = end - start + 1
                                
                                try:
                                    # Проверяем числовые имена (1-4 цифры)
                                    folders = [f for f in os.listdir(device_folder) 
                                              if os.path.isdir(os.path.join(device_folder, f)) and self.is_numeric_folder(f)]
                                    
                                    actual_count = len(folders)
                                    actual_total += actual_count
                                    
                                    status = "✅" if expected_count == actual_count else "❌"
                                    self.check_log(f"{status} {device}: {actual_count}/{expected_count}", 
                                                 "SUCCESS" if expected_count == actual_count else "ERROR", 2)
                                    
                                    if expected_count > 0 and actual_count != expected_count:
                                        attack_errors += 1
                                    
                                    # ПРОВЕРКА СОДЕРЖИМОГО ПАПОК ДЛЯ КАЖДОГО УСТРОЙСТВА
                                    self.check_log(f"🔍 Проверка содержимого {device}:", "SECTION", 2)
                                    device_content_errors = 0
                                    for folder in folders:
                                        folder_path = os.path.join(device_folder, folder)
                                        self.check_log(f"📂 Папка {folder}:", "DETAIL", 3)
                                        if not self.check_folder_content(folder_path, log_errors=True, indent=4, check_names=False):
                                            device_content_errors += 1
                                            content_errors += 1
                                    
                                    if device_content_errors == 0:
                                        self.check_log(f"✅ Содержимое {device} проверено успешно", "SUCCESS", 2)
                                    else:
                                        self.check_log(f"❌ Ошибок в содержимом {device}: {device_content_errors}", "ERROR", 2)
                                except Exception as e:
                                    self.check_log(f"❌ Ошибка доступа к папке устройства: {str(e)}", "ERROR", 2)
                                    attack_errors += 1
                    else:
                        try:
                            # Проверяем числовые имена (1-4 цифры)
                            folders = [f for f in os.listdir(attack_folder) 
                                      if os.path.isdir(os.path.join(attack_folder, f)) and self.is_numeric_folder(f)]
                            actual_total = len(folders)
                            
                            status = "✅" if structure_info['expected_total'] == actual_total else "❌"
                            self.check_log(f"{status} Всего: {actual_total}/{structure_info['expected_total']}", 
                                         "SUCCESS" if structure_info['expected_total'] == actual_total else "ERROR", 2)
                            
                            if structure_info['expected_total'] > 0 and actual_total != structure_info['expected_total']:
                                attack_errors += 1
                            
                            # ПРОВЕРКА СОДЕРЖИМОГО ПАПОК ДЛЯ ПЛОСКОЙ СТРУКТУРЫ
                            self.check_log(f"🔍 Проверка содержимого папок:", "SECTION", 2)
                            flat_content_errors = 0
                            for folder in folders:
                                folder_path = os.path.join(attack_folder, folder)
                                self.check_log(f"📂 Папка {folder}:", "DETAIL", 3)
                                if not self.check_folder_content(folder_path, log_errors=True, indent=4, check_names=False):
                                    flat_content_errors += 1
                                    content_errors += 1
                            
                            if flat_content_errors == 0:
                                self.check_log(f"✅ Содержимое папок проверено успешно", "SUCCESS", 2)
                            else:
                                self.check_log(f"❌ Ошибок в содержимом: {flat_content_errors}", "ERROR", 2)
                        except Exception as e:
                            self.check_log(f"❌ Ошибка доступа к папке атаки: {str(e)}", "ERROR", 2)
                            attack_errors += 1
                    
                    total_content_errors += content_errors
                    
                    if attack_errors == 0 and content_errors == 0:
                        self.check_log(f"✅ Атака проверена успешно", "SUCCESS", 1)
                    else:
                        error_msg = f"❌ Атака содержит ошибок: структура={attack_errors}, содержимое={content_errors}"
                        self.check_log(error_msg, "ERROR", 1)
                        total_errors += (attack_errors + content_errors)
                    
                    self.check_log("", "INFO")
                
                except Exception as e:
                    self.check_log(f"❌ Ошибка при проверке атаки: {str(e)}", "ERROR", 1)
                    total_errors += 1
            
            self.check_log("=" * 60, "HEADER")
            if total_errors == 0:
                self.check_log(f"✅ ПРОВЕРКА ID ЗАВЕРШЕНА УСПЕШНО!", "SUCCESS")
                self.check_log(f"📊 Проверено атак: {total_attacks}", "SUCCESS")
                self.check_log(f"🔍 Ошибок содержимого: {total_content_errors}", "SUCCESS")
                messagebox.showinfo("Проверка завершена", "ID проверен успешно! Ошибок не обнаружено.")
            else:
                self.check_log(f"❌ ПРОВЕРКА ID ЗАВЕРШЕНА С ОШИБКАМИ", "ERROR")
                self.check_log(f"📊 Обнаружено ошибок: {total_errors}", "ERROR")
                self.check_log(f"🔍 Ошибок содержимого: {total_content_errors}", "ERROR")
                messagebox.showwarning("Проверка завершена", 
                                     f"Обнаружены ошибки: {total_errors}\n"
                                     f"Ошибок содержимого: {total_content_errors}")
                
        except Exception as e:
            self.check_log(f"❌ Ошибка при проверке ID: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка при проверке: {str(e)}")

    def check_global(self):
        """Общая проверка проекта с проверкой содержимого папок"""
        project_folder = self.global_check_entry.get()
        
        if not project_folder:
            messagebox.showerror("Ошибка", "Выберите общую папку проекта")
            return
        
        if not os.path.exists(project_folder):
            messagebox.showerror("Ошибка", "Общая папка проекта не существует")
            return
        
        self.check_log("=" * 60, "HEADER")
        self.check_log(f"🌐 ОБЩАЯ ПРОВЕРКА ПРОЕКТА", "HEADER")
        self.check_log(f"📁 Папка: {project_folder}", "HEADER")
        self.check_log("=" * 60, "HEADER")
        
        try:
            id_folders = []
            for item in os.listdir(project_folder):
                item_path = os.path.join(project_folder, item)
                if os.path.isdir(item_path):
                    try:
                        has_attacks = any(subitem in self.attack_ranges for subitem in os.listdir(item_path))
                        if has_attacks:
                            id_folders.append(item_path)
                    except:
                        continue
            
            if not id_folders:
                self.check_log(f"❌ ОШИБКА: В проекте не найдено папок ID", "ERROR")
                return
            
            total_errors = 0
            total_ids = len(id_folders)
            total_content_errors = 0
            
            self.check_log(f"📊 Найдено ID: {total_ids}", "INFO")
            self.check_log("", "INFO")
            
            for id_folder in id_folders:
                self.check_log(f"🆔 ID: {os.path.basename(id_folder)}", "SECTION")
                
                try:
                    attack_folders = []
                    unknown_folders = []
                    
                    for item in os.listdir(id_folder):
                        item_path = os.path.join(id_folder, item)
                        if os.path.isdir(item_path):
                            if item in self.attack_ranges:
                                attack_folders.append((item, item_path))
                            else:
                                unknown_folders.append(item)
                    
                    if not attack_folders:
                        self.check_log(f"❌ В папке ID не найдено папок атак", "ERROR", 1)
                        total_errors += 1
                        continue
                    
                    self.check_log(f"📊 Количество атак: {len(attack_folders)}", "INFO", 1)
                    
                    if unknown_folders:
                        self.check_log(f"⚠️ Неизвестные папки: {', '.join(unknown_folders)}", "WARNING", 1)
                    
                    id_errors = 0
                    id_content_errors = 0
                    
                    for attack_name, attack_folder in attack_folders:
                        self.check_log(f"🎯 Атака: {attack_name}", "INFO", 2)
                        
                        try:
                            structure_info = self.check_attack_structure(attack_folder, attack_name)
                            
                            actual_total = 0
                            attack_errors = 0
                            attack_content_errors = 0
                            
                            if structure_info['has_kozen10'] or structure_info['has_kozen12']:
                                for device in ["kozen 10", "kozen 12"]:
                                    if structure_info[f'has_{device.replace(" ", "")}']:
                                        device_folder = os.path.join(attack_folder, device)
                                        
                                        if not os.path.exists(device_folder):
                                            self.check_log(f"⚠️ Папка {device} не существует", "WARNING", 3)
                                            continue
                                        
                                        expected_count = 0
                                        if attack_name in self.attack_ranges and device in self.attack_ranges[attack_name]:
                                            start, end = self.attack_ranges[attack_name][device]
                                            expected_count = end - start + 1
                                        
                                        try:
                                            # Проверяем числовые имена (1-4 цифры)
                                            folders = [f for f in os.listdir(device_folder) 
                                                      if os.path.isdir(os.path.join(device_folder, f)) and self.is_numeric_folder(f)]
                                            
                                            actual_count = len(folders)
                                            actual_total += actual_count
                                            
                                            status = "✅" if expected_count == actual_count else "❌"
                                            self.check_log(f"{status} {device}: {actual_count}/{expected_count}", 
                                                         "SUCCESS" if expected_count == actual_count else "ERROR", 3)
                                            
                                            if expected_count > 0 and actual_count != expected_count:
                                                attack_errors += 1
                                            
                                            # ПРОВЕРКА СОДЕРЖИМОГО
                                            self.check_log(f"🔍 Проверка содержимого {device}:", "SECTION", 3)
                                            device_content_errors = 0
                                            for folder in folders:
                                                folder_path = os.path.join(device_folder, folder)
                                                self.check_log(f"📂 Папка {folder}:", "DETAIL", 4)
                                                if not self.check_folder_content(folder_path, log_errors=True, indent=5, check_names=False):
                                                    device_content_errors += 1
                                                    attack_content_errors += 1
                                            
                                            if device_content_errors == 0:
                                                self.check_log(f"✅ Содержимое {device} OK", "SUCCESS", 3)
                                            else:
                                                self.check_log(f"❌ Ошибок в {device}: {device_content_errors}", "ERROR", 3)
                                        except Exception as e:
                                            self.check_log(f"❌ Ошибка доступа к папке устройства: {str(e)}", "ERROR", 3)
                                            attack_errors += 1
                            else:
                                try:
                                    # Проверяем числовые имена (1-4 цифры)
                                    folders = [f for f in os.listdir(attack_folder) 
                                              if os.path.isdir(os.path.join(attack_folder, f)) and self.is_numeric_folder(f)]
                                    actual_total = len(folders)
                                    
                                    status = "✅" if structure_info['expected_total'] == actual_total else "❌"
                                    self.check_log(f"{status} Всего: {actual_total}/{structure_info['expected_total']}", 
                                                 "SUCCESS" if structure_info['expected_total'] == actual_total else "ERROR", 3)
                                    
                                    if structure_info['expected_total'] > 0 and actual_total != structure_info['expected_total']:
                                        attack_errors += 1
                                    
                                    # ПРОВЕРКА СОДЕРЖИМОГО
                                    self.check_log(f"🔍 Проверка содержимого:", "SECTION", 3)
                                    flat_content_errors = 0
                                    for folder in folders:
                                        folder_path = os.path.join(attack_folder, folder)
                                        self.check_log(f"📂 Папка {folder}:", "DETAIL", 4)
                                        if not self.check_folder_content(folder_path, log_errors=True, indent=5, check_names=False):
                                            flat_content_errors += 1
                                            attack_content_errors += 1
                                    
                                    if flat_content_errors == 0:
                                        self.check_log(f"✅ Содержимое папок OK", "SUCCESS", 3)
                                    else:
                                        self.check_log(f"❌ Ошибок в содержимом: {flat_content_errors}", "ERROR", 3)
                                except Exception as e:
                                    self.check_log(f"❌ Ошибка доступа к папке атаки: {str(e)}", "ERROR", 3)
                                    attack_errors += 1
                            
                            id_content_errors += attack_content_errors
                            total_content_errors += attack_content_errors
                            
                            if attack_errors == 0 and attack_content_errors == 0:
                                self.check_log(f"✅ Атака проверена успешно", "SUCCESS", 3)
                            else:
                                error_msg = f"❌ Ошибки: структура={attack_errors}, содержимое={attack_content_errors}"
                                self.check_log(error_msg, "ERROR", 3)
                                id_errors += (attack_errors + attack_content_errors)
                        
                        except Exception as e:
                            self.check_log(f"❌ Ошибка при проверке атаки: {str(e)}", "ERROR", 3)
                            id_errors += 1
                    
                    total_errors += id_errors
                    
                    if id_errors == 0:
                        self.check_log(f"✅ ID проверен успешно", "SUCCESS", 1)
                    else:
                        self.check_log(f"❌ ID содержит ошибок: {id_errors}", "ERROR", 1)
                
                except Exception as e:
                    self.check_log(f"❌ Ошибка при проверке ID: {str(e)}", "ERROR", 1)
                    total_errors += 1
                
                self.check_log("", "INFO")
            
            self.check_log("=" * 60, "HEADER")
            if total_errors == 0:
                self.check_log(f"✅ ОБЩАЯ ПРОВЕРКА ЗАВЕРШЕНА УСПЕШНО!", "SUCCESS")
                self.check_log(f"📊 Проверено ID: {total_ids}", "SUCCESS")
                self.check_log(f"🔍 Ошибок содержимого: {total_content_errors}", "SUCCESS")
                messagebox.showinfo("Проверка завершена", "Проект проверен успешно! Ошибок не обнаружено.")
            else:
                self.check_log(f"❌ ОБЩАЯ ПРОВЕРКА ЗАВЕРШЕНА С ОШИБКАМИ", "ERROR")
                self.check_log(f"📊 Обнаружено ошибок: {total_errors}", "ERROR")
                self.check_log(f"📊 Проверено ID: {total_ids}", "INFO")
                self.check_log(f"🔍 Ошибок содержимого: {total_content_errors}", "ERROR")
                messagebox.showwarning("Проверка завершена", 
                                     f"Обнаружены ошибки: {total_errors}\n"
                                     f"Ошибок содержимого: {total_content_errors}\n"
                                     f"Проверено ID: {total_ids}")
                
        except Exception as e:
            self.check_log(f"❌ Ошибка при общей проверке проекта: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка при проверке: {str(e)}")

    def calculate_attack_time(self):
        """Подсчёт времени съёмки для отдельной атаки"""
        attack_folder = self.attack_check_entry.get()
        
        if not attack_folder:
            messagebox.showerror("Ошибка", "Выберите папку атаки")
            return
        
        if not os.path.exists(attack_folder):
            messagebox.showerror("Ошибка", "Папка атаки не существует")
            return
        
        self.check_log("=" * 60, "HEADER")
        self.check_log(f"⏱️ ПОДСЧЁТ ВРЕМЕНИ СЪЁМКИ АТАКИ", "HEADER")
        self.check_log(f"📁 Папка: {attack_folder}", "HEADER")
        self.check_log("=" * 60, "HEADER")
        
        try:
            attack_name = os.path.basename(attack_folder)
            id_name = os.path.basename(os.path.dirname(attack_folder))
            
            # Собираем все папки с числовыми именами в атаке
            all_folder_paths = []
            
            # Проверяем структуру атаки
            structure_info = self.check_attack_structure(attack_folder, attack_name)
            
            if structure_info['has_kozen10'] or structure_info['has_kozen12']:
                for device in ["kozen 10", "kozen 12"]:
                    if structure_info[f'has_{device.replace(" ", "")}']:
                        device_folder = os.path.join(attack_folder, device)
                        if os.path.exists(device_folder):
                            folders = [f for f in os.listdir(device_folder) 
                                      if os.path.isdir(os.path.join(device_folder, f)) and self.is_numeric_folder(f)]
                            # Добавляем полные пути к папкам
                            all_folder_paths.extend([os.path.join(device_folder, folder) for folder in folders])
                            self.check_log(f"📱 Устройство {device}: {len(folders)} папок", "INFO")
            else:
                folders = [f for f in os.listdir(attack_folder) 
                          if os.path.isdir(os.path.join(attack_folder, f)) and self.is_numeric_folder(f)]
                # Добавляем полные пути к папкам
                all_folder_paths.extend([os.path.join(attack_folder, folder) for folder in folders])
                self.check_log(f"📁 Папки в корне: {len(folders)} папок", "INFO")
            
            if not all_folder_paths:
                self.check_log(f"❌ В атаке не найдено папок для подсчёта времени", "ERROR")
                return
            
            # Вычисляем время съёмки
            shooting_time = self.calculate_shooting_time_for_folders(all_folder_paths)
            
            # Определяем дату съёмки на основе EXIF данных
            shooting_date = self.get_common_shooting_date(all_folder_paths)
            
            # Удаляем существующую запись для этого ID и атаки
            self.shooting_report_data = [item for item in self.shooting_report_data 
                                        if not (item['ID'] == id_name and item['Attack'] == attack_name)]
            
            # Добавляем новую запись
            self.shooting_report_data.append({
                'ID': id_name,
                'Attack': attack_name,
                'Date': shooting_date,
                'ShootingTime': shooting_time,
                'FolderCount': len(all_folder_paths)
            })
            
            self.check_log(f"✅ Время съёмки подсчитано успешно!", "SUCCESS")
            self.check_log(f"📊 ID: {id_name}", "INFO")
            self.check_log(f"🎯 Атака: {attack_name}", "INFO")
            self.check_log(f"⏱️ Время съёмки: {shooting_time}", "INFO")
            self.check_log(f"📅 Дата съёмки: {shooting_date}", "INFO")
            self.check_log(f"📁 Обработано папок: {len(all_folder_paths)}", "INFO")
            
            messagebox.showinfo("Подсчёт завершён", 
                              f"Время съёмки подсчитано!\n\n"
                              f"ID: {id_name}\n"
                              f"Атака: {attack_name}\n"
                              f"Время съёмки: {shooting_time}\n"
                              f"Дата съёмки: {shooting_date}\n"
                              f"Папок: {len(all_folder_paths)}")
            
        except Exception as e:
            self.check_log(f"❌ Ошибка при подсчёте времени: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка при подсчёте времени: {str(e)}")

    def calculate_id_time(self):
        """Подсчёт времени съёмки для всего ID"""
        id_folder = self.id_check_entry.get()
        
        if not id_folder:
            messagebox.showerror("Ошибка", "Выберите папку ID")
            return
        
        if not os.path.exists(id_folder):
            messagebox.showerror("Ошибка", "Папка ID не существует")
            return
        
        self.check_log("=" * 60, "HEADER")
        self.check_log(f"⏱️ ПОДСЧЁТ ВРЕМЕНИ СЪЁМКИ ID", "HEADER")
        self.check_log(f"📁 Папка: {id_folder}", "HEADER")
        self.check_log("=" * 60, "HEADER")
        
        try:
            id_name = os.path.basename(id_folder)
            
            # Находим все атаки в ID
            attack_folders = []
            for item in os.listdir(id_folder):
                item_path = os.path.join(id_folder, item)
                if os.path.isdir(item_path) and item in self.attack_ranges:
                    attack_folders.append((item, item_path))
            
            if not attack_folders:
                self.check_log(f"❌ В ID не найдено папок атак", "ERROR")
                return
            
            total_folders = 0
            total_time_seconds = 0
            
            for attack_name, attack_folder in attack_folders:
                self.check_log(f"🎯 Обработка атаки: {attack_name}", "SECTION")
                
                # Собираем все папки с числовыми именами в атаке
                all_folder_paths = []
                
                # Проверяем структуру атаки
                structure_info = self.check_attack_structure(attack_folder, attack_name)
                
                if structure_info['has_kozen10'] or structure_info['has_kozen12']:
                    for device in ["kozen 10", "kozen 12"]:
                        if structure_info[f'has_{device.replace(" ", "")}']:
                            device_folder = os.path.join(attack_folder, device)
                            if os.path.exists(device_folder):
                                folders = [f for f in os.listdir(device_folder) 
                                          if os.path.isdir(os.path.join(device_folder, f)) and self.is_numeric_folder(f)]
                                # Добавляем полные пути к папкам
                                all_folder_paths.extend([os.path.join(device_folder, folder) for folder in folders])
                else:
                    folders = [f for f in os.listdir(attack_folder) 
                              if os.path.isdir(os.path.join(attack_folder, f)) and self.is_numeric_folder(f)]
                    # Добавляем полные пути к папкам
                    all_folder_paths.extend([os.path.join(attack_folder, folder) for folder in folders])
                
                if all_folder_paths:
                    # Вычисляем время съёмки для атаки
                    shooting_time = self.calculate_shooting_time_for_folders(all_folder_paths)
                    total_folders += len(all_folder_paths)
                    
                    # Определяем дату съёмки на основе EXIF данных
                    shooting_date = self.get_common_shooting_date(all_folder_paths)
                    
                    # Преобразуем время в секунды для подсчета общего времени
                    if shooting_time != "не удалось вычислить":
                        time_parts = shooting_time.split(':')
                        if len(time_parts) == 3:
                            hours, minutes, seconds = map(int, time_parts)
                            total_time_seconds += hours * 3600 + minutes * 60 + seconds
                    
                    # Удаляем существующую запись для этого ID и атаки
                    self.shooting_report_data = [item for item in self.shooting_report_data 
                                                if not (item['ID'] == id_name and item['Attack'] == attack_name)]
                    
                    # Сохраняем данные для отчёта
                    self.shooting_report_data.append({
                        'ID': id_name,
                        'Attack': attack_name,
                        'Date': shooting_date,
                        'ShootingTime': shooting_time,
                        'FolderCount': len(all_folder_paths)
                    })
                    
                    self.check_log(f"✅ Атака {attack_name}: {shooting_time} ({len(all_folder_paths)} папок)", "SUCCESS")
                else:
                    self.check_log(f"⚠️ В атаке {attack_name} не найдено папок", "WARNING")
            
            # Вычисляем общее время для ID
            total_time_formatted = self.format_duration(total_time_seconds) if total_time_seconds > 0 else "не удалось вычислить"
            
            self.check_log("", "INFO")
            self.check_log(f"✅ Подсчёт времени для ID завершён!", "SUCCESS")
            self.check_log(f"📊 ID: {id_name}", "INFO")
            self.check_log(f"🎯 Обработано атак: {len(attack_folders)}", "INFO")
            self.check_log(f"📁 Всего папок: {total_folders}", "INFO")
            self.check_log(f"⏱️ Общее время съёмки ID: {total_time_formatted}", "INFO")
            
            messagebox.showinfo("Подсчёт завершён", 
                              f"Время съёмки подсчитано для всего ID!\n\n"
                              f"ID: {id_name}\n"
                              f"Атак: {len(attack_folders)}\n"
                              f"Всего папок: {total_folders}\n"
                              f"Общее время съёмки: {total_time_formatted}")
            
        except Exception as e:
            self.check_log(f"❌ Ошибка при подсчёте времени: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка при подсчёте времени: {str(e)}")

    def calculate_project_time(self):
        """Подсчёт времени съёмки для всего проекта"""
        project_folder = self.global_check_entry.get()
        
        if not project_folder:
            messagebox.showerror("Ошибка", "Выберите общую папку проекта")
            return
        
        if not os.path.exists(project_folder):
            messagebox.showerror("Ошибка", "Общая папка проекта не существует")
            return
        
        self.check_log("=" * 60, "HEADER")
        self.check_log(f"⏱️ ПОДСЧЁТ ВРЕМЕНИ СЪЁМКИ ПРОЕКТА", "HEADER")
        self.check_log(f"📁 Папка: {project_folder}", "HEADER")
        self.check_log("=" * 60, "HEADER")
        
        try:
            # Находим все ID в проекте
            id_folders = []
            for item in os.listdir(project_folder):
                item_path = os.path.join(project_folder, item)
                if os.path.isdir(item_path):
                    try:
                        has_attacks = any(subitem in self.attack_ranges for subitem in os.listdir(item_path))
                        if has_attacks:
                            id_folders.append(item_path)
                    except:
                        continue
            
            if not id_folders:
                self.check_log(f"❌ В проекте не найдено папок ID", "ERROR")
                return
            
            total_attacks = 0
            total_folders = 0
            total_time_seconds = 0
            
            for id_folder in id_folders:
                id_name = os.path.basename(id_folder)
                self.check_log(f"🆔 Обработка ID: {id_name}", "SECTION")
                
                # Находим все атаки в ID
                attack_folders = []
                for item in os.listdir(id_folder):
                    item_path = os.path.join(id_folder, item)
                    if os.path.isdir(item_path) and item in self.attack_ranges:
                        attack_folders.append((item, item_path))
                
                if not attack_folders:
                    self.check_log(f"⚠️ В ID {id_name} не найдено атак", "WARNING", 1)
                    continue
                
                id_attacks = 0
                id_folders_count = 0
                id_time_seconds = 0
                
                for attack_name, attack_folder in attack_folders:
                    self.check_log(f"🎯 Обработка атаки: {attack_name}", "INFO", 2)
                    
                    # Собираем все папки с числовыми именами в атаке
                    all_folder_paths = []
                    
                    # Проверяем структуру атаки
                    structure_info = self.check_attack_structure(attack_folder, attack_name)
                    
                    if structure_info['has_kozen10'] or structure_info['has_kozen12']:
                        for device in ["kozen 10", "kozen 12"]:
                            if structure_info[f'has_{device.replace(" ", "")}']:
                                device_folder = os.path.join(attack_folder, device)
                                if os.path.exists(device_folder):
                                    folders = [f for f in os.listdir(device_folder) 
                                              if os.path.isdir(os.path.join(device_folder, f)) and self.is_numeric_folder(f)]
                                    # Добавляем полные пути к папкам
                                    all_folder_paths.extend([os.path.join(device_folder, folder) for folder in folders])
                    else:
                        folders = [f for f in os.listdir(attack_folder) 
                                  if os.path.isdir(os.path.join(attack_folder, f)) and self.is_numeric_folder(f)]
                        # Добавляем полные пути к папкам
                        all_folder_paths.extend([os.path.join(attack_folder, folder) for folder in folders])
                    
                    if all_folder_paths:
                        # Вычисляем время съёмки для атаки
                        shooting_time = self.calculate_shooting_time_for_folders(all_folder_paths)
                        id_attacks += 1
                        id_folders_count += len(all_folder_paths)
                        
                        # Определяем дату съёмки на основе EXIF данных
                        shooting_date = self.get_common_shooting_date(all_folder_paths)
                        
                        # Преобразуем время в секунды для подсчета общего времени
                        if shooting_time != "не удалось вычислить":
                            time_parts = shooting_time.split(':')
                            if len(time_parts) == 3:
                                hours, minutes, seconds = map(int, time_parts)
                                attack_time_seconds = hours * 3600 + minutes * 60 + seconds
                                id_time_seconds += attack_time_seconds
                                total_time_seconds += attack_time_seconds
                        
                        # Удаляем существующую запись для этого ID и атаки
                        self.shooting_report_data = [item for item in self.shooting_report_data 
                                                    if not (item['ID'] == id_name and item['Attack'] == attack_name)]
                        
                        # Сохраняем данные для отчёта
                        self.shooting_report_data.append({
                            'ID': id_name,
                            'Attack': attack_name,
                            'Date': shooting_date,
                            'ShootingTime': shooting_time,
                            'FolderCount': len(all_folder_paths)
                        })
                        
                        self.check_log(f"✅ {attack_name}: {shooting_time} ({len(all_folder_paths)} папок)", "SUCCESS", 3)
                    else:
                        self.check_log(f"⚠️ В атаке {attack_name} не найдено папок", "WARNING", 3)
                
                total_attacks += id_attacks
                total_folders += id_folders_count
                
                # Форматируем время для ID
                id_time_formatted = self.format_duration(id_time_seconds) if id_time_seconds > 0 else "не удалось вычислить"
                
                self.check_log(f"📊 ID {id_name}: {id_attacks} атак, {id_folders_count} папок, время: {id_time_formatted}", "INFO", 1)
            
            # Форматируем общее время для проекта
            total_time_formatted = self.format_duration(total_time_seconds) if total_time_seconds > 0 else "не удалось вычислить"
            
            self.check_log("", "INFO")
            self.check_log(f"✅ Подсчёт времени для проекта завершён!", "SUCCESS")
            self.check_log(f"📊 Обработано ID: {len(id_folders)}", "INFO")
            self.check_log(f"🎯 Всего атак: {total_attacks}", "INFO")
            self.check_log(f"📁 Всего папок: {total_folders}", "INFO")
            self.check_log(f"⏱️ Общее время съёмки проекта: {total_time_formatted}", "INFO")
            
            messagebox.showinfo("Подсчёт завершён", 
                              f"Время съёмки подсчитано для всего проекта!\n\n"
                              f"ID: {len(id_folders)}\n"
                              f"Атак: {total_attacks}\n"
                              f"Всего папок: {total_folders}\n"
                              f"Общее время съёмки: {total_time_formatted}")
            
        except Exception as e:
            self.check_log(f"❌ Ошибка при подсчёте времени: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка при подсчёте времени: {str(e)}")

    def export_shooting_report(self):
        """Выгрузка отчёта о времени съёмки в Excel"""
        if not self.shooting_report_data:
            messagebox.showwarning("Предупреждение", "Нет данных для выгрузки. Сначала выполните подсчёт времени съёмки.")
            return
        
        try:
            # Создаем DataFrame из данных
            df = pd.DataFrame(self.shooting_report_data)
            
            # Генерируем имя файла с текущей датой и временем
            current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            default_filename = f"отчёт по времени съёмки атак_{current_time}.xlsx"
            
            # Сохраняем в Excel файл
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Сохранить отчёт о времени съёмки",
                initialfile=default_filename
            )
            
            if file_path:
                # Создаем Excel файл с красивым оформлением
                wb = Workbook()
                ws = wb.active
                ws.title = "Отчёт о времени съёмки"
                
                # Заголовок отчёта
                ws.merge_cells('A1:E1')
                title_cell = ws.cell(row=1, column=1, value="Отчёт о времени съёмки атак")
                title_cell.font = Font(bold=True, size=16, color="4f46e5")
                title_cell.alignment = Alignment(horizontal="center")
                
                # Информация о дате формирования
                ws.merge_cells('A2:E2')
                date_cell = ws.cell(row=2, column=1, value=f"Сформирован: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                date_cell.font = Font(italic=True, size=10, color="64748b")
                date_cell.alignment = Alignment(horizontal="center")
                
                # Пустая строка
                ws.append([])
                
                # Заголовки таблицы
                headers = ['ID', 'Атака', 'Дата съёмки', 'Время съёмки', 'Количество папок']
                ws.append(headers)
                
                # Данные
                for data in self.shooting_report_data:
                    ws.append([
                        data['ID'],
                        data['Attack'],
                        data['Date'],
                        data['ShootingTime'],
                        data.get('FolderCount', '')
                    ])
                
                # Стилизация
                # Заголовки таблицы
                thin_border = Border(left=Side(style='thin'), 
                                   right=Side(style='thin'), 
                                   top=Side(style='thin'), 
                                   bottom=Side(style='thin'))
                
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=4, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                
                # Данные таблицы
                for row in range(5, len(self.shooting_report_data) + 5):
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        if col in [4, 5]:  # Время съёмки и количество папок
                            cell.alignment = Alignment(horizontal="center")
                
                # Объединяем ячейки с одинаковыми ID
                current_id = None
                start_row = 5
                
                for row in range(5, len(self.shooting_report_data) + 5):
                    id_value = ws.cell(row=row, column=1).value
                    
                    if current_id is None:
                        current_id = id_value
                        start_row = row
                    elif id_value != current_id:
                        if start_row != row - 1:
                            ws.merge_cells(f'A{start_row}:A{row-1}')
                            # Центрируем объединенную ячейку
                            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                        current_id = id_value
                        start_row = row
                
                # Объединяем последнюю группу
                if start_row != len(self.shooting_report_data) + 4:
                    ws.merge_cells(f'A{start_row}:A{len(self.shooting_report_data) + 4}')
                    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                
                # Автоширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Сохраняем файл
                wb.save(file_path)
                
                self.check_log(f"✅ Отчёт успешно выгружен: {file_path}", "SUCCESS")
                messagebox.showinfo("Успех", f"Отчёт успешно выгружен в файл:\n{file_path}")
                
        except Exception as e:
            self.check_log(f"❌ Ошибка при выгрузке отчёта: {str(e)}", "ERROR")
            messagebox.showerror("Ошибка", f"Произошла ошибка при выгрузке отчёта: {str(e)}")

    def load_attack_data(self, event=None):
        """Загрузка данных выбранной атаки для редактирования"""
        attack = self.edit_attack_var.get()
        if attack in self.attack_ranges:
            ranges = self.attack_ranges[attack]
            
            kozen10 = ranges.get("kozen 10", (0, 0))
            kozen12 = ranges.get("kozen 12", (0, 0))
            
            self.kozen10_entry.delete(0, tk.END)
            self.kozen10_entry.insert(0, f"{kozen10[0]}-{kozen10[1]}" if kozen10 != (0,0) else "")
            
            self.kozen12_entry.delete(0, tk.END)
            self.kozen12_entry.insert(0, f"{kozen12[0]}-{kozen12[1]}" if kozen12 != (0,0) else "")
    
    def save_attack_data(self):
        """Сохранение изменений атаки"""
        attack = self.edit_attack_var.get()
        if not attack:
            messagebox.showerror("Ошибка", "Выберите атаку для редактирования")
            return
        
        try:
            kozen10_str = self.kozen10_entry.get().strip()
            kozen12_str = self.kozen12_entry.get().strip()
            
            new_ranges = {}
            
            if kozen10_str:
                start, end = map(int, kozen10_str.split('-'))
                new_ranges["kozen 10"] = (start, end)
            
            if kozen12_str:
                start, end = map(int, kozen12_str.split('-'))
                new_ranges["kozen 12"] = (start, end)
            
            if not new_ranges:
                messagebox.showerror("Ошибка", "Заполните хотя бы один диапазон")
                return
            
            self.attack_ranges[attack] = new_ranges
            self.save_attack_config()
            
            attacks = list(self.attack_ranges.keys())
            self.attack_combo['values'] = attacks
            self.edit_attack_combo['values'] = attacks
            
            self.log(f"Атака {attack} успешно сохранена", "SUCCESS")
            messagebox.showinfo("Успех", f"Атака {attack} успешно сохранена")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка сохранения: {str(e)}")
    
    def new_attack(self):
        """Создание новой атаки"""
        attack = tk.simpledialog.askstring("Новая атака", "Введите название новой атаки:")
        if attack:
            if attack in self.attack_ranges:
                messagebox.showerror("Ошибка", "Атака с таким названием уже существует")
                return
            
            self.attack_ranges[attack] = {}
            self.save_attack_config()
            
            attacks = list(self.attack_ranges.keys())
            self.attack_combo['values'] = attacks
            self.edit_attack_combo['values'] = attacks
            
            self.edit_attack_var.set(attack)
            self.load_attack_data()
            
            self.log(f"Создана новая атака: {attack}", "SUCCESS")
    
    def rename_attack(self):
        """Переименование существующей атаки"""
        old_attack = self.edit_attack_var.get()
        if not old_attack:
            messagebox.showerror("Ошибка", "Выберите атаку для переименования")
            return
        
        new_attack = tk.simpledialog.askstring("Переименование атаки", 
                                              f"Введите новое название для атаки {old_attack}:",
                                              initialvalue=old_attack)
        if new_attack:
            if new_attack in self.attack_ranges:
                messagebox.showerror("Ошибка", "Атака с таким названием уже существует")
                return
            
            attack_data = self.attack_ranges[old_attack]
            del self.attack_ranges[old_attack]
            self.attack_ranges[new_attack] = attack_data
            self.save_attack_config()
            
            attacks = list(self.attack_ranges.keys())
            self.attack_combo['values'] = attacks
            self.edit_attack_combo['values'] = attacks
            
            self.attack_var.set(new_attack)
            self.edit_attack_var.set(new_attack)
            self.load_attack_data()
            
            self.log(f"Атака переименована: {old_attack} → {new_attack}", "SUCCESS")
            messagebox.showinfo("Успех", f"Атака успешно переименована: {old_attack} → {new_attack}")
    
    def delete_attack(self):
        """Удаление атаки"""
        attack = self.edit_attack_var.get()
        if not attack:
            messagebox.showerror("Ошибка", "Выберите атаку для удаления")
            return
        
        if messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить атаку {attack}?"):
            del self.attack_ranges[attack]
            self.save_attack_config()
            
            attacks = list(self.attack_ranges.keys())
            self.attack_combo['values'] = attacks
            self.edit_attack_combo['values'] = attacks
            
            if attacks:
                self.edit_attack_var.set(attacks[0])
                self.load_attack_data()
            else:
                self.edit_attack_var.set("")
                self.kozen10_entry.delete(0, tk.END)
                self.kozen12_entry.delete(0, tk.END)
            
            self.log(f"Атака {attack} удалена", "SUCCESS")

def main():
    root = tk.Tk()
    app = ModernFolderRenamer(root)
    root.mainloop()

if __name__ == "__main__":
    main()